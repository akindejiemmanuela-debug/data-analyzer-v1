from flask import Flask, render_template, request, session, send_file, redirect, Response, jsonify, url_for
import pandas as pd
import os
from werkzeug.utils import secure_filename
import math
from openpyxl import Workbook
import sqlite3
import uuid
import json
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Image,
    Table,
    TableStyle
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import os
from insight_engine import generate_insight
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak
import base64
from io import BytesIO
from reportlab.platypus import Image
import io
from werkzeug.security import generate_password_hash, check_password_hash
import requests
import json
import numpy as np
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
    Image
)

import base64


import io
from io import BytesIO
from reportlab.platypus import Image
from reportlab.lib.utils import ImageReader



BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE_PATH = os.path.join(BASE_DIR, "database.db")

# Keep secrets outside the source code in production.
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

app = Flask("__name__")
app.secret_key = os.environ.get("SECRET_KEY", "dev-only-change-this-secret")
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10MB limit

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "static", "charts"), exist_ok=True)
ALLOWED_EXTENSIONS = {"csv", "xlsx"}


conn = sqlite3.connect(DATABASE_PATH)
cursor = conn.cursor()



def get_dataframe(workspace_id=None):

    # =========================================================
    # 1. LOAD DATA FROM SPECIFIC WORKSPACE
    # =========================================================

    if workspace_id is not None:

        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT worksheet_data, uploaded_file
            FROM workspaces
            WHERE id = ?
            AND username = ?
        """, (workspace_id, session["user"]))

        row = cursor.fetchone()

        conn.close()

        if row is None:
            return None

        worksheet_data = row[0]
        uploaded_file = row[1]

        # =====================================================
        # 2. WORKSHEET DATA
        # =====================================================

        if worksheet_data:

            try:
                worksheet = json.loads(worksheet_data)
            except (json.JSONDecodeError, TypeError):
                return None

            # Make sure worksheet contains data
            if not worksheet or len(worksheet) < 1:
                return None

            # First row = column headers
            headers = worksheet[0]

            # Remaining rows = data
            rows = worksheet[1:]

            # Remove completely empty rows
            rows = [
                row for row in rows
                if any(
                    str(cell).strip() != ""
                    for cell in row
                )
            ]

            # Make sure every row has same number
            # of columns as the headers
            cleaned_rows = []

            for row in rows:

                row = list(row)

                if len(row) < len(headers):
                    row += [""] * (len(headers) - len(row))

                elif len(row) > len(headers):
                    row = row[:len(headers)]

                cleaned_rows.append(row)

            df = pd.DataFrame(
                cleaned_rows,
                columns=headers
            )

            # Remove completely empty columns
            df = df.dropna(
                axis=1,
                how="all"
            )

            return df

        # =====================================================
        # 3. UPLOADED CSV / EXCEL
        # =====================================================

        if uploaded_file:

            if not os.path.exists(uploaded_file):
                return None

            try:

                if uploaded_file.lower().endswith(".csv"):

                    return pd.read_csv(
                        uploaded_file,
                        sep=None,
                        engine="python"
                    )

                elif uploaded_file.lower().endswith(".xlsx"):

                    return pd.read_excel(
                        uploaded_file
                    )

            except Exception as e:

                print("Error loading uploaded file:", e)

                return None

    # =========================================================
    # 4. OLD SESSION UPLOAD FALLBACK
    # =========================================================

    if "uploaded_file" in session:

        filepath = session["uploaded_file"]

        if os.path.exists(filepath):

            try:

                if filepath.lower().endswith(".csv"):

                    return pd.read_csv(
                        filepath,
                        sep=None,
                        engine="python"
                    )

                elif filepath.lower().endswith(".xlsx"):

                    return pd.read_excel(filepath)

            except Exception as e:

                print("Error loading session file:", e)

                return None

    # =========================================================
    # 5. NOTHING FOUND
    # =========================================================

    return None
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS
@app.route("/")
def home():

    return render_template("index.html")

@app.route("/analyze", methods=["GET", "POST"])
def analyze():

    if "user" not in session:
        return redirect("/login")

    # Get workspace ID from URL, form, or active session
    workspace_id = (
        request.args.get("workspace_id", type=int)
        or request.form.get("workspace_id", type=int)
        or session.get("workspace_id")
    )

    # If a file is uploaded, save it
    if "file" in request.files and request.files["file"].filename != "":

        file = request.files["file"]

        if not allowed_file(file.filename):
            return "Only CSV and Excel files are allowed."

        filename = secure_filename(file.filename)

        filepath = os.path.join(
            app.config["UPLOAD_FOLDER"],
            filename
        )

        file.save(filepath)

        # Keep the actual file available for get_dataframe()
        session["uploaded_file"] = filepath

        # Save uploaded workspace in database
        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO workspaces (
                username,
                workspace_name,
                workspace_type,
                worksheet_data,
                uploaded_file
            )
            VALUES (?, ?, ?, ?, ?)
        """, (
            session["user"],
            filename,
            "uploaded",
            None,
            filepath
        ))

        workspace_id = cursor.lastrowid

        conn.commit()
        conn.close()

        # Make this the active workspace
        session["workspace_id"] = workspace_id

    # Make sure a workspace exists
    if workspace_id is None:
        return "No workspace selected."

    # Keep the workspace active
    session["workspace_id"] = workspace_id

    # Load dataframe from worksheet or uploaded file
    df = get_dataframe(workspace_id)

    if df is None:
        return "No worksheet or uploaded file found."

    # Convert columns to numeric where possible
    # without using deprecated errors="ignore"
    for col in df.columns:
        converted = pd.to_numeric(df[col], errors="coerce")

    # Only convert if all non-empty values are numeric
        non_empty = df[col].notna() & (df[col].astype(str).str.strip() != "")

        if non_empty.any() and converted[non_empty].notna().all():
            df[col] = converted

        # Only replace the column if conversion produced
        # at least one valid numeric value
        if converted.notna().any():
            df[col] = converted

    # Dataset information
    missing_values = df.isnull().sum().sum()
    duplicates_rows = df.duplicated().sum()

    rows = df.shape[0]
    columns = df.shape[1]

    # Start displayed index from 1
    df.index = df.index + 1

    # Convert dataframe to HTML
    table = df.to_html(
        classes="table table-striped table-bordered table-hover",
        index=True
    )

    return render_template(
        "result.html",
        rows=rows,
        columns=columns,
        table=table,
        duplicates_rows=duplicates_rows,
        missing_values=missing_values,
        workspace_id=workspace_id
    )
@app.route("/groupby", methods=["GET"])
def groupby():

    if "user" not in session:
        return redirect("/login")

    # Get workspace ID from URL first
    workspace_id = request.args.get("workspace_id", type=int)

    # If URL doesn't contain it, use active workspace
    if workspace_id is None:
        workspace_id = session.get("workspace_id")

    print("GROUPBY GET WORKSPACE ID:", workspace_id)

    if workspace_id is None:
        return "Workspace ID is missing."

    # Make sure this remains the active workspace
    session["workspace_id"] = workspace_id

    df = get_dataframe(workspace_id)

    print("GROUPBY DATAFRAME:")
    print(df)

    if df is None:
        return """
        <h2>No dataset available.</h2>
        <p>Please upload a CSV/Excel file or open a workspace first.</p>
        <a href="/workspace">Go to Workspace</a>
        """

    # Remove unwanted unnamed columns
    df = df.loc[:, ~df.columns.astype(str).str.contains("^Unnamed")]

    return render_template(
        "groupby.html",
        columns=df.columns.tolist(),
        workspace_id=workspace_id
    )


@app.route("/groupby", methods=["POST"])
def groupby_result():

    if "user" not in session:
        return redirect("/login")

    # Get workspace ID from form
    workspace_id = request.form.get("workspace_id", type=int)

    # Fallback to active workspace
    if workspace_id is None:
        workspace_id = session.get("workspace_id")

    print("GROUPBY POST WORKSPACE ID:", workspace_id)

    if workspace_id is None:
        return "Workspace ID is missing."

    # Keep it active
    session["workspace_id"] = workspace_id

    # Load dataset
    df = get_dataframe(workspace_id)

    print("GROUPBY POST DATAFRAME:")
    print(df)

    if df is None:
        return """
        <h2>No dataset found.</h2>
        <p>Please open a workspace first.</p>
        <a href="/workspace">Go to Workspace</a>
        """

    # Remove unnamed columns
    df = df.loc[:, ~df.columns.astype(str).str.contains("^Unnamed")]

    group_column = request.form.get("group_column")
    value_column = request.form.get("value_column")
    operation = request.form.get("operation")

    # Validate columns
    if group_column not in df.columns:
        return "Invalid group column."

    if value_column not in df.columns:
        return "Invalid value column."

    # Numeric operations
    numeric_operations = [
        "sum",
        "mean",
        "min",
        "max",
        "std",
        "prod",
        "var"
    ]

    if operation in numeric_operations:
        df[value_column] = pd.to_numeric(
            df[value_column],
            errors="coerce"
        )

    # Perform operation
    if operation == "mean":
        result = df.groupby(group_column)[value_column].mean()

    elif operation == "sum":
        result = df.groupby(group_column)[value_column].sum()

    elif operation == "min":
        result = df.groupby(group_column)[value_column].min()

    elif operation == "max":
        result = df.groupby(group_column)[value_column].max()

    elif operation == "std":
        result = df.groupby(group_column)[value_column].std()

    elif operation == "prod":
        result = df.groupby(group_column)[value_column].prod()

    elif operation == "var":
        result = df.groupby(group_column)[value_column].var()

    elif operation == "first":
        result = df.groupby(group_column)[value_column].first()

    elif operation == "count":
        result = df.groupby(group_column)[value_column].count()

    elif operation == "last":
        result = df.groupby(group_column)[value_column].last()

    elif operation == "nunique":
        result = df.groupby(group_column)[value_column].nunique()

    else:
        return "Invalid operation."

    # Save result
    result.to_csv("groupby.csv")

    # Convert to HTML
    table = result.to_frame().to_html(
        classes="table table-bordered table-striped",
        index=True
    )

    # Generate AI insight
    report = generate_insight(
        analysis_type="groupby",
        group_column=group_column,
        value_column=value_column,
        operation=operation,
        result=result
    )

    return render_template(
        "overall.html",
        title="Group By Result",
        table=table,
        explanation=report["explanation"],
        insight=report["insight"],
        recommendation=report["recommendation"],
        learn=report["learn"],
        ai_summary=report["ai_summary"],
        back_url=f"/groupby?workspace_id={workspace_id}",
        download_url="/download/groupby.csv",
        workspace_id=workspace_id
    )
@app.route("/pivot_table")
def pivot_table():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.args.get("workspace_id", type=int)

    df = get_dataframe(workspace_id)

    if df is None:
        return """
<h2>No dataset available.</h2>
<p>Please upload a CSV/Excel file or open a worksheet first.</p>
<a href="/">Go to Home</a>
"""

    # Remove unwanted columns
    if "Unnamed: 0" in df.columns:
        df.drop(columns=["Unnamed: 0"], inplace=True)

    df = df.loc[:, ~df.columns.str.contains("^Unnamed")]

    return render_template(
        "pivot.html",
        columns=df.columns,
        workspace_id=workspace_id
    )
@app.route("/pivot_table", methods=["POST"])
def pivot_table_result():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.form.get("workspace_id", type=int)

    df = get_dataframe(workspace_id)

    if df is None:
        return "No dataset found."

    # Remove unwanted columns
    if "Unnamed: 0" in df.columns:
        df.drop(columns=["Unnamed: 0"], inplace=True)

    df = df.loc[:, ~df.columns.str.contains("^Unnamed")]

    index = request.form.get("index")
    values = request.form.get("values")
    columns = request.form.get("columns")
    aggfunc = request.form.get("aggfunc")

    # Convert selected value column to numeric
    df[values] = pd.to_numeric(df[values], errors="coerce")

    result = df.pivot_table(
        values=values,
        index=index,
        columns=columns,
        aggfunc=aggfunc
    )

    result.to_csv("pivot_table.csv")

    table = result.to_html(
        classes="table table-bordered table-striped",
        index=True
    )

    report = generate_insight(
        analysis_type="pivot_table",
        group_column=index,
        value_column=values,
        operation=aggfunc,
        result=result,
        pivot_columns=columns
    )

    return render_template(
        "overall.html",
        title="Pivot Table Result",
        table=table,
        explanation=report["explanation"],
        insight=report["insight"],
        recommendation=report["recommendation"],
        learn=report["learn"],
        ai_summary=report["ai_summary"],
        back_url=f"/pivot_table?workspace_id={workspace_id}" if workspace_id else "/pivot_table",
        download_url="/download/pivot_table.csv"
    )
@app.route("/data_cleaning")
def data_cleaning():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.args.get("workspace_id", type=int)

    if not workspace_id:
        return "Workspace ID is missing."

    df = get_dataframe(workspace_id)

    if df is None:
        return "No dataset found."

    return render_template(
        "cleaning.html",
        workspace_id=workspace_id
    )
@app.route("/cleaning", methods=["POST"])
def cleaning():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.form.get("workspace_id", type=int)

    if not workspace_id:
        return "Workspace ID is missing."

    df = get_dataframe(workspace_id)

    if df is None:
        return "No dataset found."

    operation = request.form.get("operation")

    if operation == "drop_duplicates":
        df = df.drop_duplicates()

    elif operation == "dropna":
        df = df.dropna()

    elif operation == "fillna":
        df = df.fillna(0)

    elif operation == "reset_index":
        df = df.reset_index(drop=True)

    table = df.to_html(
        classes="table table-striped table-bordered",
        index=False
    )

    return render_template(
        "result_cleaning.html",
        table=table,
        workspace_id=workspace_id
    )
@app.route("/filter")
def filter():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.args.get("workspace_id", type=int)

    df = get_dataframe(workspace_id)

    if df is None:
        return """
<h2>No dataset available.</h2>
<p>Please upload a CSV/Excel file or open a worksheet first.</p>
<a href="/">Go to Home</a>
"""

    return render_template(
        "filtering.html",
        columns=df.columns,
        workspace_id=workspace_id
    )

@app.route("/filter", methods=["POST"])
def filter_result():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.form.get("workspace_id", type=int)

    df = get_dataframe(workspace_id)

    if df is None:
        return "No dataset found."

    column = request.form.get("column")
    condition = request.form.get("condition")
    value = request.form.get("value")

    # Try converting the value to number
    try:
        value = float(value)
        df[column] = pd.to_numeric(df[column], errors="coerce")
    except:
        pass

    if condition == "==":
        result = df[df[column] == value]

    elif condition == "!=":
        result = df[df[column] != value]

    elif condition == ">":
        result = df[df[column] > value]

    elif condition == "<":
        result = df[df[column] < value]

    elif condition == ">=":
        result = df[df[column] >= value]

    else:
        result = df[df[column] <= value]

    result.to_csv("filtered.csv", index=False)

    table = result.to_html(
        classes="table table-striped table-bordered table-hover",
        index=True
    )

    return render_template(
        "result_filtering.html",
        table=table
    )

@app.route("/sorting", methods=["GET"])
def sorting():

    if "user" not in session:
        return redirect("/login")

    workspace_id = (
        request.args.get("workspace_id", type=int)
        or session.get("workspace_id")
    )

    if workspace_id:
        session["workspace_id"] = workspace_id

    df = get_dataframe(workspace_id)

    if df is None:
        return """
        <h2>No dataset available.</h2>
        <p>Please upload a CSV/Excel file or open a worksheet first.</p>
        <a href="/workspace">Go to Workspace</a>
        """

    # Remove Unnamed columns
    df = df.loc[
        :,
        ~df.columns.astype(str).str.startswith("Unnamed")
    ]

    return render_template(
        "sort.html",
        columns=df.columns,
        workspace_id=workspace_id
    )


@app.route("/sort", methods=["POST"])
def data_sorting():

    if "user" not in session:
        return redirect("/login")

    workspace_id = (
        request.form.get("workspace_id", type=int)
        or session.get("workspace_id")
    )

    if not workspace_id:
        return "Workspace ID is missing."

    df = get_dataframe(workspace_id)

    if df is None:
        return "No dataset found."

    # Remove Unnamed columns
    df = df.loc[
        :,
        ~df.columns.astype(str).str.startswith("Unnamed")
    ]

    column = request.form.get("column")
    order = request.form.get("order")

    # Make sure column exists
    if not column:
        return "Please select a column to sort."

    if column not in df.columns:
        return f"Column '{column}' was not found."

    # Try converting selected column to numeric
    converted = pd.to_numeric(
        df[column],
        errors="coerce"
    )

    # Only replace the column if it actually contains numeric data
    if converted.notna().any():
        df[column] = converted

    # Sort
    if order == "ascending":

        result = df.sort_values(
            by=column,
            ascending=True
        )

    else:

        result = df.sort_values(
            by=column,
            ascending=False
        )

    # Save sorting result
    result.to_csv(
        "sorting.csv",
        index=False
    )

    # Create preview
    table = result.to_html(
        classes="table table-striped table-bordered table-hover",
        index=False
    )

    return render_template(
        "result_sort.html",
        table=table,
        workspace_id=workspace_id
    )
@app.route("/statistics")
def statistics():

    if "user" not in session:
        return redirect("/login")

    # Get workspace ID from URL first,
    # otherwise use the active workspace
    workspace_id = (
        request.args.get("workspace_id", type=int)
        or session.get("workspace_id")
    )

    if workspace_id is None:
        return """
        <h2>Workspace ID is missing.</h2>
        <p>Please open a workspace first.</p>
        <a href="/workspace">Go to Workspace</a>
        """

    # Make this the active workspace
    session["workspace_id"] = workspace_id

    print("STATISTICS GET WORKSPACE ID:", workspace_id)

    # Load dataset
    df = get_dataframe(workspace_id)

    if df is None:
        return """
        <h2>No dataset available.</h2>
        <p>Please upload a CSV/Excel file or open a worksheet first.</p>
        <a href="/workspace">Go to Workspace</a>
        """

    # Remove all Unnamed columns
    df.columns = df.columns.astype(str).str.strip()

    df = df.loc[
        :,
        ~df.columns.str.match(
            r"^Unnamed",
            case=False,
            na=False
        )
    ]

    return render_template(
        "statistics.html",
        columns=df.columns,
        workspace_id=workspace_id
    )
@app.route("/statistics", methods=["POST"])
def statistics_result():

    if "user" not in session:
        return redirect("/login")

    # ==========================
    # GET WORKSPACE ID
    # ==========================

    workspace_id = (
        request.form.get("workspace_id", type=int)
        or session.get("workspace_id")
    )

    if workspace_id is None:
        return """
        <h2>Workspace ID is missing.</h2>
        <p>Please open a workspace first.</p>
        <a href="/workspace">Go to Workspace</a>
        """

    # Make this the active workspace
    session["workspace_id"] = workspace_id

    print("STATISTICS WORKSPACE ID:", workspace_id)

    # ==========================
    # LOAD DATASET
    # ==========================

    df = get_dataframe(workspace_id)

    print("STATISTICS DATAFRAME:")
    print(df)

    if df is None:
        return """
        <h2>No dataset found.</h2>
        <p>Please upload a CSV/Excel file or open a workspace first.</p>
        <a href="/workspace">Go to Workspace</a>
        """

    # ==========================
    # REMOVE UNNAMED COLUMNS
    # ==========================

    df.columns = df.columns.astype(str).str.strip()

    df = df.loc[
        :,
        ~df.columns.str.match(
            r"^Unnamed",
            case=False,
            na=False
        )
    ]

    # ==========================
    # CONVERT NUMERIC COLUMNS
    # ==========================

    for col in df.columns:

        converted = pd.to_numeric(
            df[col],
            errors="coerce"
        )

        # Only replace the column if it actually
        # contains numeric values
        if converted.notna().any():
            df[col] = converted

    # ==========================
    # GET OPERATION
    # ==========================

    operation = request.form.get("operation")

    if not operation:
        return "No statistics operation selected."

    # ==========================
    # STATISTICS
    # ==========================

    if operation == "describe":

        numeric = df.    select_dtypes(include="number")
        categorical = df.select_dtypes(exclude="number")

        numeric_result = pd.DataFrame()
        categorical_result = pd.DataFrame()

    # ==========================
    # NUMERIC STATISTICS
    # ==========================

        if not numeric.empty:

            numeric_result = numeric.describe().T

            numeric_result.insert(
            0,
                "Column",
                numeric_result.index
            )

            numeric_result.reset_index(
                drop=True,
                inplace=True
            )

    # ==========================
    # CATEGORICAL STATISTICS
    # ==========================

        if not categorical.empty:

            categorical_result = categorical.describe().T

            categorical_result.insert(
                0,
                "Column",
                categorical_result.index
            )

            categorical_result.reset_index(
                drop=True,
                inplace=True
            )

    # Combine both
        result = pd.concat(
            [
                numeric_result,
                categorical_result
            ],
            ignore_index=True,
            sort=False
        )

    # Remove NaN from display
        result = result.fillna("")
    elif operation == "mean":

        result = df.mean(
            numeric_only=True
        ).to_frame(
            name="Mean"
        )

    elif operation == "median":

        result = df.median(
            numeric_only=True
        ).to_frame(
            name="Median"
        )

    elif operation == "mode":

        mode_summary = []

        for col in df.columns:

            values = df[col].mode()
            counts = df[col].value_counts()

            if counts.empty:

                mode = "No data"
                frequency = 0

            elif counts.max() == 1:

                mode = "No repeated value"
                frequency = 1

            elif len(values) == 1:

                mode = values.iloc[0]
                frequency = counts.loc[mode]

            else:

                mode = ", ".join(
                    map(str, values.tolist())
                )

                frequency = ", ".join(
                    str(counts.loc[v])
                    for v in values
                )

            mode_summary.append({

                "Column": col,

                "Most Frequent Value": mode,

                "Frequency": frequency

            })

        result = pd.DataFrame(
            mode_summary
        )

    elif operation == "sum":

        result = df.sum(
            numeric_only=True
        ).to_frame(
            name="Sum"
        )

    elif operation == "count":

        result = df.count().to_frame(
            name="Count"
        )

    elif operation == "min":

        result = df.min(
            numeric_only=True
        ).to_frame(
            name="Minimum"
        )

    elif operation == "max":

        result = df.max(
            numeric_only=True
        ).to_frame(
            name="Maximum"
        )

    elif operation == "std":

        result = df.std(
            numeric_only=True
        ).to_frame(
            name="Standard Deviation"
        )

    elif operation == "var":

        result = df.var(
            numeric_only=True
        ).to_frame(
            name="Variance"
        )

    elif operation == "skew":

        result = df.skew(
            numeric_only=True
        ).to_frame(
            name="Skewness"
        )

    elif operation == "kurtosis":

        result = df.kurt(
            numeric_only=True
        ).to_frame(
            name="Kurtosis"
        )

    else:

        return "Invalid statistics operation."

    # ==========================
    # SAVE RESULT
    # ==========================

    result.to_csv(
        "statistics.csv",
        index=False
    )

    # ==========================
    # CREATE HTML TABLE
    # ==========================

    table = result.to_html(
        classes="table table-striped table-bordered table-hover",
        index=True
    )

    # ==========================
    # GENERATE AI INSIGHT
    # ==========================

    report = generate_insight(
        analysis_type="statistics",
        operation=operation,
        result=result
    )

    # ==========================
    # SHOW RESULT
    # ==========================

    return render_template(
        "overall.html",

        title="Statistics Result",

        table=table,

        explanation=report["explanation"],

        insight=report["insight"],

        recommendation=report["recommendation"],

        learn=report["learn"],

        ai_summary=report["ai_summary"],

        back_url=(
            f"/statistics?workspace_id={workspace_id}"
        ),

        download_url="/download/statistics.csv"
    )
@app.route("/analysis")
def analysis():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.args.get("workspace_id", type=int)

    df = get_dataframe(workspace_id)

    if df is None:
        return """
<h2>No dataset available.</h2>
<p>Please upload a CSV/Excel file or open a worksheet first.</p>
<a href="/">Go to Home</a>
"""

    # Remove unwanted columns
    if "Unnamed: 0" in df.columns:
        df.drop(columns=["Unnamed: 0"], inplace=True)

    df = df.loc[:, ~df.columns.str.contains("^Unnamed")]

    return render_template(
        "analysis.html",
        columns=df.columns,
        workspace_id=workspace_id
    )

@app.route("/analysis", methods=["POST"])
def analysis_result():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.form.get("workspace_id", type=int)

    df = get_dataframe(workspace_id)

    if df is None:
        return "No dataset found."

    # -----------------------------
    # Remove unwanted columns
    # -----------------------------

    df = df.loc[:, ~df.columns.astype(str).str.contains(
        r"^Unnamed",
        case=False,
        regex=True
    )]

    column = request.form.get("column")
    operation = request.form.get("operation")

    if not column or column not in df.columns:
        return "Invalid column selected."

    # -----------------------------
    # Try numeric conversion
    # -----------------------------

    converted = pd.to_numeric(df[column], errors="coerce")

    # Use numeric version only if the column actually contains numbers
    if converted.notna().any():
        analysis_series = converted
        is_numeric = True
    else:
        analysis_series = df[column]
        is_numeric = False

    # -----------------------------
    # Basic information
    # -----------------------------

    total_values = len(analysis_series)
    missing_values = analysis_series.isna().sum()
    unique_values = analysis_series.nunique(dropna=True)

    # -----------------------------
    # Perform analysis
    # -----------------------------

    if operation == "count":

        result_value = analysis_series.count()
        result_name = "Count"

    elif operation == "sum":

        if not is_numeric:
            return "Sum can only be performed on a numeric column."

        result_value = analysis_series.sum()
        result_name = "Sum"

    elif operation == "mean":

        if not is_numeric:
            return "Mean can only be performed on a numeric column."

        result_value = analysis_series.mean()
        result_name = "Mean"

    elif operation == "median":

        if not is_numeric:
            return "Median can only be performed on a numeric column."

        result_value = analysis_series.median()
        result_name = "Median"

    elif operation == "mode":

        values = analysis_series.mode()

        if len(values) == 0:

            result_value = "No mode"

            frequency = 0

        else:

            result_value = ", ".join(
                str(value) for value in values.tolist()
            )

            counts = analysis_series.value_counts()

            frequency = ", ".join(
                str(counts.get(value, 0))
                for value in values.tolist()
            )

        result_name = "Most Frequent Value"

    elif operation == "min":

        if not is_numeric:
            return "Minimum can only be calculated for numeric data."

        result_value = analysis_series.min()
        result_name = "Minimum"

    elif operation == "max":

        if not is_numeric:
            return "Maximum can only be calculated for numeric data."

        result_value = analysis_series.max()
        result_name = "Maximum"

    elif operation == "std":

        if not is_numeric:
            return "Standard deviation requires numeric data."

        result_value = analysis_series.std()
        result_name = "Standard Deviation"

    elif operation == "var":

        if not is_numeric:
            return "Variance requires numeric data."

        result_value = analysis_series.var()
        result_name = "Variance"

    elif operation == "skew":

        if not is_numeric:
            return "Skewness requires numeric data."

        result_value = analysis_series.skew()
        result_name = "Skewness"

    elif operation == "kurt":

        if not is_numeric:
            return "Kurtosis requires numeric data."

        result_value = analysis_series.kurt()
        result_name = "Kurtosis"

    elif operation == "nunique":

        result_value = analysis_series.nunique(
            dropna=True
        )

        result_name = "Unique Values"

    else:

        return "Invalid analysis operation."

    # -----------------------------
    # Round numeric results
    # -----------------------------

    if isinstance(result_value, (float, np.floating)):

        result_value = round(float(result_value), 2)

    # -----------------------------
    # Create result table
    # -----------------------------

    if operation == "mode":

        result = pd.DataFrame({
            "Most Frequent Value": [result_value],
            "Frequency": [frequency]
        })

    else:

        result = pd.DataFrame({
            "Metric": [result_name],
            "Value": [result_value]
        })

    result.to_csv(
        "analysis.csv",
        index=False
    )

    table = result.to_html(
        classes="table table-striped table-bordered table-hover",
        index=False
    )

    # -----------------------------
    # Numeric statistics
    # -----------------------------

    numeric_stats = {}

    if is_numeric:

        numeric_stats = {
            "mean": round(float(analysis_series.mean()), 2),
            "median": round(float(analysis_series.median()), 2),
            "minimum": round(float(analysis_series.min()), 2),
            "maximum": round(float(analysis_series.max()), 2),
            "std": round(float(analysis_series.std()), 2)
        }

    # -----------------------------
    # Generate explanation
    # -----------------------------

    if operation == "mean":

        explanation = (
            f"The average value of '{column}' is "
            f"{result_value}."
        )

    elif operation == "median":

        explanation = (
            f"The middle value of '{column}' is "
            f"{result_value}."
        )

    elif operation == "sum":

        explanation = (
            f"The total value of '{column}' is "
            f"{result_value}."
        )

    elif operation == "max":

        explanation = (
            f"The highest value in '{column}' is "
            f"{result_value}."
        )

    elif operation == "min":

        explanation = (
            f"The lowest value in '{column}' is "
            f"{result_value}."
        )

    elif operation == "count":

        explanation = (
            f"'{column}' contains {result_value} "
            f"non-empty values."
        )

    elif operation == "nunique":

        explanation = (
            f"'{column}' contains {result_value} "
            f"unique values."
        )

    elif operation == "mode":

        explanation = (
            f"The most frequently occurring value "
            f"in '{column}' is {result_value}."
        )

    else:

        explanation = (
            f"The {operation} of '{column}' is "
            f"{result_value}."
        )

    # -----------------------------
    # Insight
    # -----------------------------

    if missing_values > 0:

        insight = (
            f"The column contains {missing_values} "
            f"missing values out of {total_values} records."
        )

    else:

        insight = (
            f"The column contains no missing values "
            f"and has {unique_values} unique values."
        )

    # -----------------------------
    # Recommendation
    # -----------------------------

    if is_numeric:

        recommendation = (
            "Compare the mean and median to understand "
            "the distribution of the data. You can also "
            "use a histogram or box plot for deeper analysis."
        )

    else:

        recommendation = (
            "For this categorical column, consider using "
            "frequency analysis, Group By, or a bar chart."
        )

    # -----------------------------
    # Render page
    # -----------------------------

    return render_template(
        "result_analysis.html",

        table=table,

        column=column,

        operation=operation,

        result_value=result_value,

        result_name=result_name,

        total_values=total_values,

        missing_values=missing_values,

        unique_values=unique_values,

        is_numeric=is_numeric,

        numeric_stats=numeric_stats,

        explanation=explanation,

        insight=insight,

        recommendation=recommendation,

        workspace_id=workspace_id
    )



def add_chart(title, image_data):

    if not image_data:
        print(title, "NOT RECEIVED")
        return

    try:

        print(title, "RECEIVED")

        # Remove:
        # data:image/png;base64,
        if "," in image_data:
            image_data = image_data.split(",", 1)[1]

        image_bytes = base64.b64decode(image_data)

        image_stream = io.BytesIO(image_bytes)

        img = Image(
            image_stream,
            width=500,
            height=280
        )

        story.append(
            Paragraph(
                f"<b>{title}</b>",
                styles["Heading2"]
            )
        )

        story.append(img)

        story.append(
            Spacer(1, 20)
        )

    except Exception as e:

        print(
            f"{title} ERROR:",
            str(e)
        )
@app.route("/correlation")
def correlation():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.args.get("workspace_id", type=int)

    df = get_dataframe(workspace_id)

    if df is None:
        return """
<h2>No dataset available.</h2>
<p>Please upload a CSV/Excel file or open a worksheet first.</p>
<a href="/">Go to Home</a>
"""

    return render_template(
        "correlation.html",
        workspace_id=workspace_id
    )

@app.route("/correlation", methods=["POST"])
def correlation_result():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.form.get(
        "workspace_id",
        type=int
    ) or session.get("workspace_id")

    df = get_dataframe(workspace_id)

    if df is None:
        return "No dataset found."

    # -----------------------------
    # Remove unwanted columns
    # -----------------------------

    df = df.loc[
        :,
        ~df.columns.astype(str).str.contains(
            r"^Unnamed",
            case=False,
            regex=True
        )
    ]

    # -----------------------------
    # Select numeric columns only
    # -----------------------------

    numeric_df = df.select_dtypes(
        include="number"
    )

    # -----------------------------
    # Check if enough numeric columns
    # -----------------------------

    if numeric_df.shape[1] < 2:

        return """
        <div style="font-family: Arial; padding: 30px;">

            <h2>⚠️ Not Enough Numeric Columns</h2>

            <p>
                Correlation analysis requires at least
                two numeric columns.
            </p>

            <p>
                Your dataset currently has fewer than
                two usable numeric columns.
            </p>

            <a href="/dashboard">
                Go Back
            </a>

        </div>
        """

    # -----------------------------
    # Calculate correlation
    # -----------------------------

    result = numeric_df.corr()

    # -----------------------------
    # Round values
    # -----------------------------

    result = result.round(3)

    # -----------------------------
    # Save result
    # -----------------------------

    result.to_csv(
        "correlation.csv"
    )

    # -----------------------------
    # Convert to HTML
    # -----------------------------

    table = result.to_html(
        classes="""
        table
        table-bordered
        table-striped
        table-hover
        text-center
        """,
        index=True
    )

    # -----------------------------
    # Find strongest relationship
    # -----------------------------

    strongest_pair = None
    strongest_value = 0

    columns = result.columns

    for i in range(len(columns)):

        for j in range(i + 1, len(columns)):

            value = result.iloc[i, j]

            if pd.notna(value):

                if abs(value) > abs(strongest_value):

                    strongest_value = value

                    strongest_pair = (
                        columns[i],
                        columns[j]
                    )

    # -----------------------------
    # Generate insight
    # -----------------------------

    if strongest_pair:

        col1, col2 = strongest_pair

        if strongest_value > 0:

            relationship = "positive"

        elif strongest_value < 0:

            relationship = "negative"

        else:

            relationship = "no"

        insight = (
            f"The strongest relationship is between "
            f"'{col1}' and '{col2}', with a correlation "
            f"of {strongest_value:.3f}. "
            f"This indicates a {relationship} relationship."
        )

    else:

        insight = (
            "No meaningful relationship could be "
            "identified between the numeric columns."
        )

    # -----------------------------
    # Render result
    # -----------------------------

    return render_template(
        "result_correlation.html",

        table=table,

        workspace_id=workspace_id,

        numeric_columns=list(
            numeric_df.columns
        ),

        strongest_pair=strongest_pair,

        strongest_value=round(
            strongest_value,
            3
        ),

        insight=insight
    )
@app.route("/charts", methods=["GET"])
def chart():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.args.get(
        "workspace_id",
        type=int
    )

    # If workspace_id wasn't supplied in the URL,
    # try the current session workspace
    if workspace_id is None:
        workspace_id = session.get("workspace_id")

    if workspace_id is None:
        return """
        <h2>No workspace selected.</h2>
        <p>Please open a workspace first.</p>
        <a href="/">Go Home</a>
        """

    df = get_dataframe(workspace_id)

    if df is None:
        return """
        <h2>No dataset available.</h2>
        <p>Please upload a CSV/Excel file or open a worksheet first.</p>
        <a href="/">Go Home</a>
        """

    # Convert numeric-looking columns
    for col in df.columns:
        df[col] = pd.to_numeric(
            df[col],
            errors="ignore"
        )

    columns = df.columns.tolist()

    numeric_columns = df.select_dtypes(
        include="number"
    ).columns.tolist()

    return render_template(
        "chart.html",

        columns=columns,

        numeric_columns=numeric_columns,

        workspace_id=workspace_id
    )

@app.route("/charts", methods=["POST"])
def chart_result():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.form.get(
        "workspace_id",
        type=int
    )

    if workspace_id is None:
        workspace_id = session.get("workspace_id")

    df = get_dataframe(workspace_id)

    if df is None:
        return "No dataset found."

    x = request.form.get("x")
    y = request.form.get("y")
    chart_type = request.form.get("chart")

    # Validate Y column
    if not y or y not in df.columns:
        return "Invalid Y-axis column."

    # Convert Y to numeric
    df[y] = pd.to_numeric(
        df[y],
        errors="coerce"
    )

    # Remove empty Y values
    df = df.dropna(
        subset=[y]
    )

    # -------------------------
    # HISTOGRAM
    # -------------------------

    if chart_type == "histogram":

        labels = df[y].tolist()

        values = df[y].tolist()

    # -------------------------
    # BOX PLOT
    # -------------------------

    elif chart_type == "box":

        labels = ["Data"]

        values = df[y].tolist()

    # -------------------------
    # NORMAL CHARTS
    # -------------------------

    else:

        if not x or x not in df.columns:
            return "Invalid X-axis column."

        labels = df[x].astype(str).tolist()

        values = df[y].tolist()

    # Table
    if x and x in df.columns:

        table_data = df[
            [x, y]
        ].to_dict(
            orient="records"
        )

    else:

        table_data = df[
            [y]
        ].to_dict(
            orient="records"
        )

    # Save chart information
    session["labels"] = labels
    session["values"] = values
    session["chart"] = chart_type
    session["x"] = x
    session["y"] = y
    session["workspace_id"] = workspace_id

    # Statistics
    rows = len(df)

    total = round(
        df[y].sum(),
        2
    )

    average = round(
        df[y].mean(),
        2
    )

    maximum = round(
        df[y].max(),
        2
    )

    minimum = round(
        df[y].min(),
        2
    )

    return render_template(
        "result_chart.html",

        labels=labels,

        values=values,

        chart=chart_type,

        x=x,

        y=y,

        rows=rows,

        total=total,

        average=average,

        maximum=maximum,

        minimum=minimum,

        table_data=table_data,

        workspace_id=workspace_id
    )
@app.route("/chart_data")
def chart_data():

    return {
        "labels": session.get(
            "labels",
            []
        ),

        "values": session.get(
            "values",
            []
        ),

        "chart": session.get(
            "chart",
            ""
        ),

        "x": session.get(
            "x",
            ""
        ),

        "y": session.get(
            "y",
            ""
        ),

        "workspace_id": session.get(
            "workspace_id"
        )
    }
@app.route("/dashboard")
def dashboard():

    if "user" not in session:
        return redirect("/login")

    # Get workspace ID from URL or active session
    workspace_id = (
        request.args.get("workspace_id", type=int)
        or session.get("workspace_id")
    )

    if workspace_id is None:
        return "No workspace selected."

    # Make this the active workspace
    session["workspace_id"] = workspace_id

    # Load dataset
    df = get_dataframe(workspace_id)

    if df is None:
        return render_template(
            "dashboard.html",
            rows=0,
            columns=0,
            missing=0,
            duplicates=0,
            preview="",
            chart_labels=[],
            chart_values=[],
            workspace_id=workspace_id
        )

    # Remove automatically generated unnamed columns
    df = df.loc[
        :,
        ~df.columns.astype(str).str.contains("^Unnamed")
    ]

    # Dataset statistics
    rows = df.shape[0]
    columns = df.shape[1]

    missing = int(df.isnull().sum().sum())
    duplicates = int(df.duplicated().sum())

    # Dataset preview
    preview = df.head(10).to_html(
        classes="table table-striped table-bordered table-hover",
        index=False
    )

    # Find numeric columns
    numeric = df.select_dtypes(include="number")

    chart_labels = []
    chart_values = []

    if not numeric.empty:

        # Use the first numeric column
        column = numeric.columns[0]

        chart_labels = list(range(1, len(df) + 1))

        chart_values = (
            numeric[column]
            .fillna(0)
            .tolist()
        )

    return render_template(
        "dashboard.html",
        rows=rows,
        columns=columns,
        missing=missing,
        duplicates=duplicates,
        preview=preview,
        chart_labels=chart_labels,
        chart_values=chart_values,
        workspace_id=workspace_id
    )

@app.route("/download")
def download():

    workspace_id = request.args.get("workspace_id") or session.get("workspace_id")

    df = get_dataframe(workspace_id)

    if df is None:
        return "No worksheet or uploaded file found."

    filename = session.get("filename")

    if filename is None:
        filename = "Worksheet Data"

    wb = Workbook()
    summary = wb.active
    summary.title = "Dashboard Summary"
    summary.append(["DATA ANALYTICS REPORT"])
    summary.append([])
    from datetime import datetime

    summary.append(["Generated On", datetime.now().strftime("%d-%m-%Y %H:%M")])
    summary.append(["Dataset Name", filename])
    summary.append(["Rows", len(df)])
    summary.append(["Columns", len(df.columns)])
    summary.append(["Missing Values", df.isnull().sum().sum()])
    summary.append(["Duplicate Rows", df.duplicated().sum()])
    summary.append([])
    
    summary.append(["TOP KPI", "VALUE"])

    summary.append(["Total Rows", len(df)])
    summary.append(["Total Columns", len(df.columns)])

    numeric = df.select_dtypes(include="number")

    if not numeric.empty:

        first = numeric.columns[0]

        summary.append(["Average " + first, round(df[first].mean(),2)])
        summary.append(["Maximum " + first, df[first].max()])
        summary.append(["Minimum " + first, df[first].min()])
    for cell in summary["A"]:
        cell.font = Font(bold=True)
    # Report Title
    summary["A1"].font = Font(size=18, bold=True, color="FFFFFF")

    summary["A1"].fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
    )

    summary["A1"].alignment =         Alignment(horizontal="center")

    summary.merge_cells("A1:B1")    
    
    for column_cells in summary.columns:

        length = max(
            len(str(cell.value))
            if cell.value is not None else 0
            for cell in column_cells
        )

        summary.column_dimensions[
    get_column_letter(column_cells[0].column)
    ].width = length + 5
    
    
    
    
    ws = wb.create_sheet("Original Data")

    ws.append(df.columns.tolist())

    for row in df.values.tolist():
        ws.append(row)

    # ======================
    # Sheet 2 - Statistics
    # ======================

    stat = wb.create_sheet("Statistics")

    stat.append(["Metric", "Value"])

    stat.append(["Rows", len(df)])
    stat.append(["Columns", len(df.columns)])
    stat.append(["Missing Values", df.isnull().sum().sum()])
    stat.append(["Duplicate Rows", df.duplicated().sum()])

    # ======================
    # Sheet 3 - Missing Values
    # ======================
    # ======================
# Sheet 3 - KPI Summary
# ======================

    kpi_sheet = wb.create_sheet("KPI Summary")

    kpi_sheet.append(["Metric", "Value"])

    kpi = []

# General KPIs
    kpi.append(["Total Rows", len(df)])
    kpi.append(["Total Columns", len(df.columns)])
    kpi.append(["Missing Values", df.isnull().sum().sum()])
    kpi.append(["Duplicate Rows", df.duplicated().sum()])

# Numeric Columns
    numeric_columns = df.select_dtypes(include="number").columns

    for col in numeric_columns:
        kpi.append([f"Total {col}", round(df[col].sum(), 2)])
        kpi.append([f"Average {col}", round(df[col].mean(), 2)])
        kpi.append([f"Maximum {col}", df[col].max()])
        kpi.append([f"Minimum {col}", df[col].min()])
        kpi.append([f"Median {col}", round(df[col].median(), 2)])
        kpi.append([f"Standard Deviation ({col})", round(df[col].std(), 2)])

# Text Columns
    text_columns = df.select_dtypes(include="object").columns

    for col in text_columns:
        kpi.append([f"Unique {col}", df[col].nunique()])
        kpi.append([f"Most Common {col}", df[col].mode().iloc[0]])
        kpi.append([f"Least Common {col}", df[col].value_counts().idxmin()])

# Write KPIs to Excel
    for row in kpi:
        kpi_sheet.append(row)
    miss = wb.create_sheet("Missing Values")

    miss.append(["Column", "Missing"])

    for col in df.columns:
        miss.append([col, df[col].isnull().sum()])
    
    if os.path.exists("groupby.csv"):

        group_df = pd.read_csv("groupby.csv")

        group_sheet = wb.create_sheet("Group By")

        group_sheet.append(group_df.columns.tolist())

        for row in group_df.values.tolist():
            group_sheet.append(row)
    if os.path.exists("pivot.csv"):

        pivot_df = pd.read_csv("pivot.csv")

        pivot_sheet = wb.create_sheet("Pivot Table")

        pivot_sheet.append(pivot_df.columns.tolist())

        for row in pivot_df.values.tolist():
            pivot_sheet.append(row)
        
    if os.path.exists("filtered.csv"):

        filter_df = pd.read_csv("filtered.csv")

        filter_sheet = wb.create_sheet("Filtered Data")

        filter_sheet.append(filter_df.columns.tolist())

        for row in filter_df.values.tolist():
            filter_sheet.append(row)        
    
    
    if os.path.exists("sorting.csv"):
    
        sort_df = pd.read_csv("sorting.csv")
        sort_sheet = wb.create_sheet("Sorted Data")

        sort_sheet.append(sort_df.columns.tolist())

        for row in sort_df.values.tolist():
            sort_sheet.append(row)
    if os.path.exists("cleaning.csv"):
        clean_df = pd.read_csv("cleaning.csv")
        clean_sheet = wb.create_sheet("Cleaned Data")

        clean_sheet.append(clean_df.columns.tolist())

        for row in clean_df.values.tolist():
            clean_sheet.append(row)
 
    if os.path.exists("edited.csv"):

        edit_df = pd.read_csv("edited.csv")

        edit_sheet = wb.create_sheet("Edited Data")

        edit_sheet.append(edit_df.columns.tolist())

        for row in edit_df.values.tolist():
            edit_sheet.append(row)       
    if os.path.exists("correlation.csv"):

        corr_df = pd.read_csv("correlation.csv")

        corr_sheet = wb.create_sheet("Correlation")

        corr_sheet.append(corr_df.columns.tolist())

        for row in corr_df.values.tolist():
            corr_sheet.append(row)   
    if os.path.exists("analysis.csv"):

        analysis_df = pd.read_csv("analysis.csv")

        analysis_sheet = wb.create_sheet("Analysis")

        analysis_sheet.append(analysis_df.columns.tolist())

        
        for row in analysis_df.values.tolist():
            analysis_sheet.append(row)    
     
      
     

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
)

    header_font = Font(
        color="FFFFFF",
        bold=True
)

    for sheet in wb.worksheets:

        for cell in sheet[1]:

            cell.fill = header_fill

            cell.font = header_font   
    
    for sheet in wb.worksheets:

        for column_cells in sheet.columns:

            length = max(
                len(str(cell.value))
                if cell.value is not None else 0
                for cell in column_cells
            )

            sheet.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = length + 5    
    
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"    
    
    
    
    
    report_name = "analysis_report.xlsx"

    wb.save(report_name)

    
    
    return send_file(
        report_name,
        as_attachment=True
    )
    
@app.route("/preview")
def preview():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.args.get("workspace_id") or session.get("workspace_id")

    df = get_dataframe(workspace_id)

    if df is None:
        return "No worksheet or uploaded file found."

    page = int(request.args.get("page", 1))
    per_page = 10

    total_pages = math.ceil(len(df) / per_page)

    start = (page - 1) * per_page
    end = start + per_page

    table = df.iloc[start:end].to_html(
        classes="table table-striped",
        index=False
    )

    return render_template(
        "preview.html",
        table=table,
        page=page,
        total_pages=total_pages
    )        

@app.route("/edit")
def edit():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.args.get("workspace_id") or session.get("workspace_id")

    df = get_dataframe(workspace_id)

    if df is None:
        return "No worksheet or uploaded file found."

# Remove Unnamed columns
    df = df.loc[
        :,
    ~df.columns.astype(str).str.strip().str.startswith("Unnamed")
]

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT DISTINCT system_name
        FROM grading_systems
        WHERE username = ?
        ORDER BY system_name
    """, (session["user"],))

    grading_systems = cursor.fetchall()

    conn.close()

    return render_template(

        "edit.html",

        columns=df.columns,

        grading_systems=grading_systems

    )
                               
@app.route("/edit", methods=["POST"])
def edit_result():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.form.get(
        "workspace_id",
        type=int
    ) or session.get("workspace_id")

    if workspace_id is None:
        return "Workspace ID is missing."

    df = get_dataframe(workspace_id)

    if df is None:
        return "No worksheet or uploaded file found."

    # ==========================================
    # REMOVE UNNAMED COLUMNS
    # ==========================================

    df = df.loc[
        :,
        ~df.columns.astype(str).str.strip().str.startswith("Unnamed")
    ]

    # ==========================================
    # GET FORM VALUES
    # ==========================================

    operation = request.form.get("operation", "").strip()

    column = request.form.get("column", "").strip()

    value1 = request.form.get("value1", "").strip()

    value2 = request.form.get("value2", "").strip()

    grading_system = request.form.get(
        "grading_system",
        ""
    ).strip()

    # ==========================================
    # CHECK SELECTED COLUMN
    # ==========================================

    column_operations = [
        "rename",
        "delete",
        "duplicate",
        "datatype",
        "replace",
        "fillna",
        "trim",
        "uppercase",
        "lowercase",
        "titlecase",
        "percentage",
        "passfail",
        "rank",
        "grade"
    ]

    if operation in column_operations:

        if not column:
            return "Please select a column."

        if column not in df.columns:

            return (
                f"Column '{column}' does not exist. "
                f"Available columns: "
                f"{', '.join(map(str, df.columns))}"
            )

    # ==========================================
    # COLUMN OPERATIONS
    # ==========================================

    if operation == "rename":

        if not value1:
            return "Please enter a new column name."

        if value1 in df.columns and value1 != column:
            return f"A column named '{value1}' already exists."

        df.rename(
            columns={
                column: value1
            },
            inplace=True
        )

    elif operation == "delete":

        df.drop(
            columns=[column],
            inplace=True
        )

    elif operation == "duplicate":

        new_name = value1

        if new_name == "":
            new_name = column + "_copy"

        if new_name in df.columns:
            return f"A column named '{new_name}' already exists."

        df[new_name] = df[column]

    elif operation == "datatype":

        if value1 == "text":

            df[column] = df[column].astype(str)

        elif value1 == "integer":

            df[column] = pd.to_numeric(
                df[column],
                errors="coerce"
            ).fillna(0).astype(int)

        elif value1 == "decimal":

            df[column] = pd.to_numeric(
                df[column],
                errors="coerce"
            )

        elif value1 == "date":

            df[column] = pd.to_datetime(
                df[column],
                errors="coerce"
            )

    # ==========================================
    # DATA CLEANING
    # ==========================================

    elif operation == "replace":

        df[column] = df[column].replace(
            value1,
            value2
        )

    elif operation == "fillna":

        df[column] = df[column].fillna(value1)

    elif operation == "removeduplicates":

        df.drop_duplicates(
            inplace=True
        )

    elif operation == "removeblank":

        df.dropna(
            inplace=True
        )

    elif operation == "trim":

        df[column] = (
            df[column]
            .astype(str)
            .str.strip()
        )

    elif operation == "uppercase":

        df[column] = (
            df[column]
            .astype(str)
            .str.upper()
        )

    elif operation == "lowercase":

        df[column] = (
            df[column]
            .astype(str)
            .str.lower()
        )

    elif operation == "titlecase":

        df[column] = (
            df[column]
            .astype(str)
            .str.title()
        )

    # ==========================================
    # CREATE NEW COLUMNS
    # ==========================================

    elif operation == "newcolumn":

        if not value1:
            return "Please enter a new column name."

        if value1 in df.columns:
            return f"Column '{value1}' already exists."

        df[value1] = value2

    elif operation == "formula":

        if not value1:
            return "Please enter a new column name."

        if not value2:
            return "Please enter a formula."

        try:

            df[value1] = df.eval(value2)

        except Exception as e:

            return f"Formula error: {str(e)}"

    elif operation == "percentage":

        if not value2:
            value2 = column + "_percentage"

        try:

            total = float(value1)

            df[value2] = (
                pd.to_numeric(
                    df[column],
                    errors="coerce"
                ) / total
            ) * 100

        except ValueError:

            return "Total must be a valid number."

    elif operation == "passfail":

        if not value2:
            value2 = column + "_status"

        try:

            pass_mark = float(value1)

        except ValueError:

            return "Pass mark must be a valid number."

        numeric_values = pd.to_numeric(
            df[column],
            errors="coerce"
        )

        df[value2] = numeric_values.apply(
            lambda score:
            "Pass"
            if pd.notna(score) and score >= pass_mark
            else "Fail"
        )

    elif operation == "rank":

        if not value2:
            value2 = column + "_rank"

        order = value1.lower()

        numeric_values = pd.to_numeric(
            df[column],
            errors="coerce"
        )

        if order == "lowest":

            df[value2] = numeric_values.rank(
                ascending=True,
                method="dense"
            )

        else:

            df[value2] = numeric_values.rank(
                ascending=False,
                method="dense"
            )

    elif operation == "grade":

        if not grading_system:
            return "Please select a grading system."

        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                min_score,
                max_score,
                grade
            FROM grading_systems
            WHERE username = ?
            AND system_name = ?
            ORDER BY max_score DESC
        """, (
            session["user"],
            grading_system
        ))

        rules = cursor.fetchall()

        conn.close()

        if not rules:
            return "Grading system not found."

        def assign_grade(score):

            try:
                score = float(score)
            except (ValueError, TypeError):
                return ""

            for minimum, maximum, grade in rules:

                if minimum <= score <= maximum:
                    return grade

            return ""

        new_column = value2

        if new_column == "":
            new_column = "Grade"

        df[new_column] = df[column].apply(
            assign_grade
        )

    else:

        return "Unknown edit operation."

    # ==========================================
    # REMOVE UNNAMED COLUMNS AGAIN
    # ==========================================

    df = df.loc[
        :,
        ~df.columns.astype(str).str.strip().str.startswith("Unnamed")
    ]

    # ==========================================
    # SAVE DATASET
    # ==========================================

    worksheet_data = (
        [df.columns.tolist()]
        + df.fillna("").values.tolist()
    )

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    # Check workspace type
    cursor.execute("""
        SELECT workspace_type
        FROM workspaces
        WHERE id = ?
        AND username = ?
    """, (
        workspace_id,
        session["user"]
    ))

    workspace = cursor.fetchone()

    if not workspace:
        conn.close()
        return "Workspace not found."

    workspace_type = workspace[0]

    # ==========================================
    # WORKSHEET WORKSPACE
    # ==========================================

    if workspace_type != "uploaded":

        cursor.execute("""
            UPDATE workspaces
            SET worksheet_data = ?
            WHERE id = ?
            AND username = ?
        """, (
            json.dumps(worksheet_data),
            workspace_id,
            session["user"]
        ))

    # ==========================================
    # UPLOADED WORKSPACE
    # ==========================================

    else:

        cursor.execute("""
            SELECT uploaded_file
            FROM workspaces
            WHERE id = ?
            AND username = ?
        """, (
            workspace_id,
            session["user"]
        ))

        file_row = cursor.fetchone()

        if file_row and file_row[0]:

            filepath = file_row[0]

            # Save edited dataframe back to the original file
            if filepath.lower().endswith(".csv"):

                df.to_csv(
                    filepath,
                    index=False
                )

            elif filepath.lower().endswith(".xlsx"):

                df.to_excel(
                    filepath,
                    index=False
                )

    conn.commit()
    conn.close()

    # ==========================================
    # EXPORT COPY
    # ==========================================

    df.to_csv(
        "edited.csv",
        index=False
    )

    # ==========================================
    # PREVIEW
    # ==========================================

    table = df.to_html(
        classes=(
            "table table-striped "
            "table-bordered table-hover"
        ),
        index=False
    )

    # ==========================================
    # SUMMARY
    # ==========================================

    rows = len(df)

    columns = len(df.columns)

    summary = f"""
    The dataset now contains
    {rows} rows and {columns} columns.

    The selected edit operation
    '{operation}'
    was completed successfully.
    """

    return render_template(
        "result_edit.html",
        table=table,
        summary=summary,
        workspace_id=workspace_id,
        download_url=url_for(
            "download",
            workspace_id=workspace_id
        )
    )
@app.route("/search", methods=["POST"])
def search():

    if "user" not in session:
        return redirect("/login")

    workspace_id = session.get("workspace_id")

    df = get_dataframe(workspace_id)

    if df is None:
        return """
        <h2>No dataset found.</h2>
        <a href="/">Go Home</a>
        """

    keyword = request.form.get("search", "").strip()

    if keyword == "":
        result = df

    else:
        result = df[
            df.astype(str)
              .apply(
                  lambda row: row.str.contains(
                      keyword,
                      case=False,
                      na=False
                  ).any(),
                  axis=1
              )
        ]

    table = result.to_html(
        classes="table table-striped table-hover",
        index=False
    )

    return render_template(
        "search_result.html",
        table=table,
        keyword=keyword,
        total_results=len(result)
    )
 
 
def create_users_table():

    conn = sqlite3.connect(DATABASE_PATH)

    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS users(

        id INTEGER PRIMARY KEY AUTOINCREMENT,

        fullname TEXT NOT NULL,

        email TEXT UNIQUE NOT NULL,

        username TEXT UNIQUE NOT NULL,

        password TEXT NOT NULL

    )
    """)

    conn.commit()

    conn.close()



                                 
def create_saved_charts_table():

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS saved_charts(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        chart_name TEXT,
        chart_type TEXT,
        filename TEXT
    )
    """)

    conn.commit()
    conn.close()



def create_tables():
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS shared_charts(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        share_id TEXT UNIQUE,
        labels TEXT,
        chart_values TEXT,
        chart_type TEXT
    )
    """)

    conn.commit()
    conn.close()


def create_dasboard_charts():

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS saved_dashboards(

        id INTEGER PRIMARY KEY AUTOINCREMENT,

        username TEXT,

        dashboard_name TEXT,

        filename TEXT,

        x_column TEXT,

        y_column TEXT,

        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP

    )
    """)

    conn.commit()
    conn.close()


def create_grading_system_table():

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS grading_systems(

        id INTEGER PRIMARY KEY AUTOINCREMENT,

        username TEXT,

        system_name TEXT,

        min_score REAL,

        max_score REAL,

        grade TEXT

    )
    """)

    conn.commit()
    conn.close()

def create_datasets_table():

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""

    CREATE TABLE IF NOT EXISTS datasets(

        id INTEGER PRIMARY KEY AUTOINCREMENT,

        username TEXT,

        dataset_name TEXT,

        table_name TEXT,

        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP

    )

    """)

    conn.commit()
    conn.close()

def create_workspaces_table():

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS workspaces(

        id INTEGER PRIMARY KEY AUTOINCREMENT,

        username TEXT,

        workspace_name TEXT,

        workspace_type TEXT,

        worksheet_data TEXT,

        uploaded_file TEXT,

        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP

    )
    """)

    conn.commit()
    conn.close()

create_datasets_table()


create_grading_system_table()

create_dasboard_charts()
create_users_table()                                       
  
create_saved_charts_table()
create_tables()
create_workspaces_table()


@app.route("/register")
def register():
    return render_template("register.html")

@app.route("/register", methods=["POST"])
def register_post():

    fullname = request.form.get("fullname").strip()
    email = request.form.get("email").strip()
    username = request.form.get("username").strip()
    password = request.form.get("password", "").strip()
    confirm = request.form.get("confirm", "").strip()

    if not password or not confirm:
        return "Password fields cannot be empty."

    if password != confirm:
        return "Passwords do not match."

    # 🔐 ADD THIS
    from werkzeug.security import generate_password_hash
    hashed_password = generate_password_hash(password)

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    try:
        cursor.execute(
            """
            INSERT INTO users(fullname, email, username, password)
            VALUES (?, ?, ?, ?)
            """,
            (fullname, email, username, hashed_password)
        )

        conn.commit()

    except sqlite3.IntegrityError:
        conn.close()
        return "Username or Email already exists."

    conn.close()

    return redirect("/login")                      

@app.route("/grading/<int:workspace_id>")
def grading(workspace_id):

    if "user" not in session:
        return redirect("/login")

    df = get_dataframe(workspace_id)

    if df is None:
        return "No dataset found."

    return render_template(
        "grading.html",
        workspace_id=workspace_id
    )
    
@app.route("/save_grading", methods=["POST"])
def save_grading():

    if "user" not in session:
        return redirect("/login")

    workspace_id = request.form.get("workspace_id", type=int)

    system_name = request.form.get("system_name")

    minimum = request.form.getlist("minimum")
    maximum = request.form.getlist("maximum")
    grades = request.form.getlist("grade")

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute(
        """
        DELETE FROM grading_systems
        WHERE username=?
        AND workspace_id=?
        AND system_name=?
        """,
        (
            session["user"],
            workspace_id,
            system_name
        )
    )

    for mn, mx, gd in zip(minimum, maximum, grades):

        if not mn.strip() or not mx.strip() or not gd.strip():
            continue

        cursor.execute(
            """
            INSERT INTO grading_systems
            (
                username,
                workspace_id,
                system_name,
                min_score,
                max_score,
                grade
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                session["user"],
                workspace_id,
                system_name,
                float(mn),
                float(mx),
                gd
            )
        )

    conn.commit()
    conn.close()

    return redirect(
        url_for("grading", workspace_id=workspace_id)
    )
    
@app.route("/delete_workspace/<int:id>", methods=["POST"])
def delete_workspace(id):

    if "user" not in session:
        return redirect("/login")

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT workspace_type, uploaded_file
        FROM workspaces
        WHERE id = ?
        AND username = ?
    """, (id, session["user"]))

    workspace = cursor.fetchone()

    if workspace is None:
        conn.close()
        return "Workspace not found."

    workspace_type, uploaded_file = workspace

    cursor.execute("""
        DELETE FROM workspaces
        WHERE id = ?
        AND username = ?
    """, (id, session["user"]))

    conn.commit()
    conn.close()

    # Remove the uploaded file from disk too, if this workspace had one
    if workspace_type == "uploaded" and uploaded_file:
        try:
            if os.path.exists(uploaded_file):
                os.remove(uploaded_file)
        except OSError:
            pass

    # Clear it as the active workspace if it was selected
    if session.get("workspace_id") == id:
        session.pop("workspace_id", None)

    return redirect("/workspace")    
@app.route("/login")
def login():
    return render_template("login.html")


@app.route("/login", methods=["POST"])
def login_post():

    username = request.form.get("username").strip()
    password = request.form.get("password").strip()

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    # ✅ STEP 1: get user ONLY by username
    cursor.execute(
        "SELECT * FROM users WHERE username=?",
        (username,)
    )

    user = cursor.fetchone()
    conn.close()

    # ❌ user not found
    if not user:
        return "Invalid Username or Password."

    # ✅ STEP 2: check hashed password
    if check_password_hash(user[4], password):

        session["user"] = user[1]      # fullname or username (depends on table)
        session["user_id"] = user[0]   # correct ID from DB

        return redirect("/workspace")

    return "Invalid Username or Password."                                                                              

@app.route("/logout")
def logout():

    session.pop("user", None)

    return redirect("/login")




@app.route("/ask_ai", methods=["POST"])
def ask_ai():

    if "user" not in session:
        return redirect("/login")

    workspace_id = session.get("workspace_id")

    df = get_dataframe(workspace_id)

    if df is None:
        return """
        <h2>No dataset found.</h2>
        <p>Please upload a CSV/Excel file or open a worksheet.</p>
        <a href="/workspace">Go Home</a>
        """

    # Get user's question
    question = request.form.get("question", "").strip()
    question_lower = question.lower()

    answer = "Sorry, I couldn't understand your question."

    # -----------------------------
    # Built-in dataset questions
    # -----------------------------

    if "row" in question_lower:
        answer = f"Your dataset contains {len(df)} rows."

    elif "list columns" in question_lower or "column names" in question_lower:
        answer = "Columns: " + ", ".join(map(str, df.columns))

    elif "numeric columns" in question_lower:
        numeric = df.select_dtypes(include="number")

        if numeric.empty:
            answer = "There are no numeric columns."
        else:
            answer = "Numeric columns: " + ", ".join(map(str, numeric.columns))

    elif "text columns" in question_lower:
        text = df.select_dtypes(include="object")

        if text.empty:
            answer = "There are no text columns."
        else:
            answer = "Text columns: " + ", ".join(map(str, text.columns))

    elif "column" in question_lower:
        answer = f"Your dataset contains {len(df.columns)} columns."

    elif "shape" in question_lower:
        answer = (
            f"The dataset has {df.shape[0]} rows "
            f"and {df.shape[1]} columns."
        )

    elif "missing" in question_lower:

        missing = df.isnull().sum().sum()

        answer = (
            f"There are {missing} missing values "
            f"in the dataset."
        )

    elif "most missing" in question_lower:

        missing = df.isnull().sum()

        if missing.empty:
            answer = "No columns were found."
        else:
            col = missing.idxmax()

            answer = (
                f"'{col}' has the highest number of "
                f"missing values ({missing.max()})."
            )

    elif "duplicate" in question_lower:

        duplicates = df.duplicated().sum()

        answer = (
            f"There are {duplicates} duplicate rows."
        )

    elif "highest average" in question_lower or "average" in question_lower:

        numeric = df.select_dtypes(include="number")

        if numeric.empty:

            answer = "There are no numeric columns."

        else:

            col = numeric.mean().idxmax()
            value = numeric.mean().max()

            answer = (
                f"'{col}' has the highest average value "
                f"({value:.2f})."
            )

    elif (
        "maximum" in question_lower
        or "highest value" in question_lower
    ):

        numeric = df.select_dtypes(include="number")

        if numeric.empty:

            answer = "No numeric columns found."

        else:

            col = numeric.max().idxmax()
            value = numeric.max().max()

            answer = (
                f"The highest value is {value} "
                f"in the '{col}' column."
            )

    elif (
        "minimum" in question_lower
        or "lowest value" in question_lower
    ):

        numeric = df.select_dtypes(include="number")

        if numeric.empty:

            answer = "No numeric columns found."

        else:

            col = numeric.min().idxmin()
            value = numeric.min().min()

            answer = (
                f"The lowest value is {value} "
                f"in the '{col}' column."
            )

    elif "chart" in question_lower:

        answer = (
            "I recommend using bar charts for categories, "
            "line charts for trends, pie charts for proportions, "
            "histograms for numeric distributions, and "
            "scatter plots for relationships."
        )

    elif "clean" in question_lower:

        answer = (
            "Remove duplicate rows, fill or remove missing values, "
            "check data types, and remove invalid records "
            "before analysis."
        )

    elif "recommendation" in question_lower:

        answer = (
            "I recommend checking missing values, removing "
            "duplicates, using correlation analysis, creating "
            "charts, and reviewing summary statistics."
        )

    # -----------------------------
    # Gemini fallback
    # -----------------------------

    if answer == "Sorry, I couldn't understand your question.":

        url = (
            "https://generativelanguage.googleapis.com/"
            "v1beta/models/gemini-2.5-flash:generateContent"
            f"?key={GEMINI_API_KEY}"
        )

        headers = {
            "Content-Type": "application/json"
        }

        # Dataset context
        if len(df) <= 100:

            dataset_preview = df.to_string(index=False)

        else:

            dataset_preview = (
                "Dataset Summary:\n\n"
                f"Rows: {len(df)}\n"
                f"Columns: {len(df.columns)}\n\n"
                "Column Names:\n"
                f"{', '.join(map(str, df.columns))}\n\n"
                "Missing Values:\n"
                f"{df.isnull().sum().to_string()}\n\n"
                "Summary Statistics:\n"
                f"{df.describe(include='all').to_string()}\n\n"
                "First 20 Rows:\n"
                f"{df.head(20).to_string(index=False)}"
            )

        data = {
            "contents": [
                {
                    "parts": [
                        {
                            "text": f"""
You are an expert data analyst.

Below is the dataset currently opened by the user:

{dataset_preview}

The user asked:

{question}

Answer ONLY based on this dataset.

If the answer cannot be determined from the dataset,
clearly say so.

Give clear and concise explanations.
"""
                        }
                    ]
                }
            ]
        }

        try:

            response = requests.post(
                url,
                headers=headers,
                json=data,
                timeout=30
            )

            if response.status_code == 200:

                result = response.json()

                answer = (
                    result["candidates"][0]
                    ["content"]["parts"][0]["text"]
                )

            else:

                answer = (
                    "Gemini AI could not answer your question."
                )

        except requests.RequestException:

            answer = (
                "Unable to connect to Gemini AI. "
                "Please check your internet connection."
            )

    return render_template(
        "ask_ai.html",
        question=question,
        answer=answer
    )



@app.route("/workspace")
def workspace():

    if "user" not in session:
        return redirect("/login")

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            id,
            workspace_name,
            workspace_type
        FROM workspaces
        WHERE username = ?
        ORDER BY id DESC
    """, (session["user"],))

    workspaces = cursor.fetchall()

    conn.close()

    return render_template(
        "workspace.html",
        workspaces=workspaces
    )
    
@app.route("/open_workspace/<int:id>")
def open_workspace(id):

    if "user" not in session:
        return redirect("/login")

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT workspace_type
        FROM workspaces
        WHERE id = ?
        AND username = ?
    """, (id, session["user"]))

    workspace = cursor.fetchone()
    conn.close()

    if workspace is None:
        return "Workspace not found."

    workspace_type = workspace[0]

    print("OPEN WORKSPACE ID:", id)
    print("WORKSPACE TYPE:", workspace_type)

    session["workspace_id"] = id

    # Worksheet workspace types
    if workspace_type in ["worksheet", "school"]:
        return redirect(
            url_for("worksheet", workspace_id=id)
        )

    # Uploaded file workspace
    elif workspace_type == "uploaded":
        return redirect(
            url_for("analyze", workspace_id=id)
        )

    return f"Unknown workspace type: {workspace_type}"
@app.route("/save_dashboard", methods=["POST"])
def save_dashboard():

    if "user" not in session:
        return redirect("/login")

    dashboard_name = request.form.get("dashboard_name")

    workspace_id = session.get("workspace_id")

    if workspace_id is None:
        return "No workspace selected."

    x_column = request.form.get("x_column")
    y_column = request.form.get("y_column")

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO saved_dashboards (
            username,
            dashboard_name,
            workspace_id,
            x_column,
            y_column
        )
        VALUES (?, ?, ?, ?, ?)
    """, (
        session["user"],
        dashboard_name,
        workspace_id,
        x_column,
        y_column
    ))

    conn.commit()
    conn.close()

    return redirect("/dashboard_charts")




@app.route("/saved_charts")
def saved_charts():

    if "user" not in session:
        return redirect("/login")

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
    SELECT id, chart_name, chart_type
    FROM saved_charts
    WHERE user_id=?
    ORDER BY id DESC
    """, (session["user_id"],))

    charts = cursor.fetchall()

    conn.close()

    return render_template(
        "saved_charts.html",
        charts=charts
    )

@app.route("/ai")
def ai():

    if "user" not in session:
        return redirect("/login")

    workspace_id = session.get("workspace_id")

    df = get_dataframe(workspace_id)

    if df is None:
        return """
        <h2>No dataset found.</h2>
        <p>Please upload a CSV/Excel file or open a worksheet first.</p>
        <a href="/workspace">Go Back</a>
        """

    insights = []
    recommendations = []

    # Dataset size
    insights.append(
        f"This dataset contains {len(df)} rows and {len(df.columns)} columns."
    )

    # Missing values
    missing = df.isnull().sum().sum()

    if missing == 0:
        insights.append("No missing values were found.")
    else:
        insights.append(
            f"The dataset contains {missing} missing values."
        )

        missing_column = df.isnull().sum().idxmax()
        missing_count = df.isnull().sum().max()

        insights.append(
            f"The column '{missing_column}' has the most missing values ({missing_count})."
        )

    # Duplicate rows
    duplicates = df.duplicated().sum()

    if duplicates == 0:
        insights.append("No duplicate rows were found.")
    else:
        insights.append(
            f"The dataset contains {duplicates} duplicate rows."
        )

    # Highest average numeric column
    numeric = df.select_dtypes(include="number")

    if not numeric.empty:

        highest = numeric.mean().idxmax()
        value = numeric.mean().max()

        insights.append(
            f"The column '{highest}' has the highest average value ({value:.2f})."
        )
    # Most common text value
    text = df.select_dtypes(include="object")

    if not text.empty:
        first = text.columns[0]
        mode = df[first].mode()

    if not mode.empty:
        common = mode.iloc[0]
        insights.append(
        f"The most common value in '{first}' is '{common}'."
    )
       
       

    # Recommendations
    if missing > 0:
        recommendations.append(
            "Clean missing values before further analysis."
        )

    if duplicates > 0:
        recommendations.append(
            "Remove duplicate rows to improve data quality."
        )

    if len(numeric.columns) > 1:
        recommendations.append(
            "Use correlation analysis to identify relationships between numeric columns."
        )

    if len(text.columns) > 0:
        recommendations.append(
            "Create bar charts for categorical columns."
        )

    if len(numeric.columns) > 0:
        recommendations.append(
            "Create histograms or box plots to understand numeric distributions."
        )

    if not recommendations:
        recommendations.append(
            "Your dataset appears clean and is ready for advanced analysis."
        )

    return render_template(
        "ai_insight.html",
        insights=insights,
        recommendations=recommendations
    )



@app.route("/export_csv")
def export_csv():

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([x, y])

    for i in range(len(labels)):
        writer.writerow([labels[i], values[i]])

    output.seek(0)

    return Response(
        output,
        mimetype="text/csv",
        headers={
            "Content-Disposition":
            "attachment; filename=chart_data.csv"
        }
    )


@app.route("/share_chart")
def share_chart():

    share_id = uuid.uuid4().hex[:8]

    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO shared_charts
        (share_id, labels, values, chart_type)
        VALUES (?, ?, ?, ?)
    """,
    (
        share_id,
        json.dumps(labels),
        json.dumps(values),
        chart
    ))

    conn.commit()
    conn.close()

    return f"Share this link: http://127.0.0.1:5000/chart/{share_id}"

@app.route("/chart/<share_id>")
def shared_chart(share_id):

    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()

    cur.execute(
        "SELECT labels, values, chart_type FROM shared_charts WHERE share_id=?",
        (share_id,)
    )

    row = cur.fetchone()

    conn.close()

    if not row:
        return "Chart not found"

    labels = json.loads(row[0])
    values = json.loads(row[1])
    chart = row[2]

    return render_template(
        "dashboard.html",
        labels=labels,
        values=values,
        chart=chart
    )

@app.route("/drilldown")
def drilldown():

    if "user" not in session:
        return redirect("/login")

    workspace_id = session.get("workspace_id")

    df = get_dataframe(workspace_id)

    if df is None:
        return """
        <h2>No dataset found.</h2>
        <p>Please upload a CSV/Excel file or open a worksheet first.</p>
        <a href="/workspace">Go Back</a>
        """

    category = request.args.get("category")
    x_column = request.args.get("x_column")
    y_column = request.args.get("y_column")

    # Remove unwanted columns
    df = df.loc[:, ~df.columns.str.contains("^Unnamed")]

    # Filter by clicked category
    filtered = df[df[x_column].astype(str) == str(category)]

    return render_template(
        "drilldown.html",
        rows=filtered.to_dict(orient="records"),
        columns=filtered.columns.tolist(),
        category=category,
        x_column=x_column,
        y_column=y_column
    )


@app.route("/kpi")
def kpi():

    if "user" not in session:
        return redirect("/login")

    workspace_id = session.get("workspace_id")

    df = get_dataframe(workspace_id)

    if df is None:
        return """
        <h2>No dataset found.</h2>
        <p>Please upload a CSV/Excel file or open a worksheet first.</p>
        <a href="/workspace">Go Back</a>
        """
      
    # 👇 ADD THE KPI CODE HERE
    kpi = []

# General KPIs
    kpi.append(["Total Rows", len(df)])
    kpi.append(["Total Columns", len(df.columns)])
    kpi.append(["Missing Values", df.isnull().sum().sum()])
    kpi.append(["Duplicate Rows", df.duplicated().sum()])
    numeric_columns = df.select_dtypes(include="number").columns

    for col in numeric_columns:

        kpi.append([f"Total {col}", round(df[col].sum(),2)])
        kpi.append([f"Average {col}", round(df[col].mean(),2)])
        kpi.append([f"Maximum {col}", df[col].max()])
        kpi.append([f"Minimum {col}", df[col].min()])
        kpi.append([f"Median {col}", round(df[col].median(),2)])
        kpi.append([f"Standard Deviation ({col})", round(df[col].std(),2)])
        
    text_columns = df.select_dtypes(include="object").columns

    for col in text_columns:

        kpi.append([f"Unique {col}", df[col].nunique()])

        mode = df[col].mode()

        if not mode.empty:
            most_common = mode.iloc[0]
        else:
            most_common = "N/A"

        value_counts = df[col].value_counts()

        if not value_counts.empty:
            least_common = value_counts.idxmin()
        else:
            least_common = "N/A"

        kpi.append([f"Most Common {col}", most_common])
        kpi.append([f"Least Common {col}", least_common])

    kpi_df = pd.DataFrame(kpi, columns=["Metric", "Value"])

    return render_template(
        "kpi.html",
        tables=kpi_df.to_html(
            index=False,
            classes="table table-striped table-bordered table-hover"
        )
    )


@app.route("/dashboard_charts",methods=["GET","POST"])
def dashboard_charts():


    workspace_id = request.args.get("workspace_id", type=int)

    if workspace_id is None:
        workspace_id = session.get("workspace_id")

    if workspace_id is not None:
        session["workspace_id"] = workspace_id
    df = get_dataframe(workspace_id)

    if df is None:
        return """
        <h2>No dataset found.</h2>
        <p>Please upload a CSV/Excel file or open a worksheet first.</p>
        <a href="/workspace">Go Back</a>
        """
        
    if "Unnamed: 0" in df.columns:
        df.drop(columns=["Unnamed: 0"], inplace=True)
    df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
    df = df.loc[:, ~df.columns.duplicated()]    
    numeric_columns = df.select_dtypes(include="number").columns.tolist()

    text_columns = df.select_dtypes(include="object").columns.tolist()

    date_columns = []

    for col in df.columns:

        converted = pd.to_datetime(df[col], errors="coerce")

        if converted.notna().sum() > 0:
            date_columns.append(col)
    recommendations = []
    # ==========================
# KPI Cards
# ==========================

    if len(numeric_columns) > 0:
        recommendations.append("KPI Cards")

# ==========================
# Bar Chart
# ==========================

    if len(text_columns) >= 1 and         len(numeric_columns) >= 1:
        recommendations.append("Bar Chart")

# ==========================
# Pie Chart
# ==========================

    if len(text_columns) >= 1 and len(numeric_columns) >= 1:
        recommendations.append("Pie Chart")

# ==========================
# Line Chart
# ==========================

    if len(date_columns) >= 1 and len(numeric_columns) >= 1:
        recommendations.append("Line Chart")

# ==========================
# Histogram
# ==========================

    if len(numeric_columns) >= 1:
        recommendations.append("Histogram")

# ==========================
# Scatter Plot
# ==========================

    if len(numeric_columns) >= 2:
        recommendations.append("Scatter Plot")

# ==========================
# Correlation Heatmap
# ==========================

    if len(numeric_columns) >= 2:
        recommendations.append("Correlation Heatmap")    

    # ==========================
# Select Columns
# ==========================

    if request.method == "POST":

        x_column = request.form.get("x_column")
        y_column = request.form.get("y_column")

    else:

        x_column = request.args.get("x_column")
        y_column = request.args.get("y_column")

        if not x_column:
            if len(date_columns) > 0:
                x_column = date_columns[0]
            elif len(text_columns) > 0:
                x_column = text_columns[0]
            else:
                x_column = numeric_columns[0]

    if not y_column:
        y_column = numeric_columns[0]
    dashboard = {}

    dashboard["rows"] = len(df)
    dashboard["columns"] = len(df.columns)
    dashboard["missing"] = df.isnull().sum().sum()
    dashboard["duplicates"] = df.duplicated().sum()

    if y_column:

        dashboard["total"] = round(df[y_column].sum(),2)

        dashboard["average"] = round(df[y_column].mean(),2)

        dashboard["maximum"] = df[y_column].max()

        dashboard["minimum"] = df[y_column].min()
    
    top5 = df.sort_values(
        by=y_column,
        ascending=False
    ).head().to_html(
        index=False,
        classes="table table-striped table-hover table-sm align-middle"

    )  

    insights = []

    insights.append(f"The dataset contains {len(df)} rows.")

    insights.append(f"The dataset contains {len(df.columns)} columns.")

    insights.append(f"There are {df.isnull().sum().sum()} missing values.")

    insights.append(f"There are {df.duplicated().sum()} duplicate rows.")   
    
    for col in numeric_columns:

        insights.append(
            f"{col} has an average of {round(df[col].mean(),2)}."
        )

        insights.append(
            f"The highest {col} is {df[col].max()}."
        )

        insights.append(
            f"The lowest {col} is {df[col].min()}."
        )    
    
    for col in text_columns:

        insights.append(
            f"{col} has {df[col].nunique()} unique values."
        )

        mode = df[col].mode()

        if not mode.empty:
            insights.append(
            f"The most common {col} is '{mode.iloc[0]}'."
        )
    if x_column == y_column:

        top10 = (
            df[[y_column]]
            .sort_values(by=y_column, ascending=False)
            .head(10)
            .to_html(
                index=False,
                classes="table table-striped table-hover table-sm align-middle"

            )
        )

        bottom10 = (
            df[[y_column]]
            .sort_values(by=y_column, ascending=True)
            .head(10)
            .to_html(
                index=False,
                classes="table table-striped table-hover"
                )
            )

    else:

        top10 = (
            df[[x_column, y_column]]
            .sort_values(by=y_column, ascending=False)
            .head(10)
            .to_html(
                index=False,
                classes="table table-striped table-hover"
            )
        )

        bottom10 = (
            df[[x_column, y_column]]
            .sort_values(by=y_column, ascending=True)
            .head(10)
            .to_html(
                index=False,
                classes="table table-striped table-hover table-sm align-middle"

            )
        )
    summary = []

    summary.append(f"Detected {len(numeric_columns)} numeric column(s).")
    summary.append(f"Detected {len(text_columns)} text column(s).")
    summary.append(f"Detected {len(date_columns)} date column(s).")
    summary.append(f"Generated {len(recommendations)} dashboard component(s).")    
    bar_title = f"{y_column} Analysis"
    pie_title = f"{y_column} Distribution"
    line_title = f"{y_column} Trend"
    hist_title = f"{y_column} Histogram"
    scatter_title = f"{y_column} Relationship"    
    if x_column not in df.columns or y_column not in df.columns:
        return "Selected columns do not exist."    
    
    labels = df[x_column].astype(str).tolist()

    values = df[y_column].fillna(0).tolist()
    
    return render_template(

        "dashboard_charts.html",

        dashboard=dashboard,

        recommendations=recommendations,
        labels=labels,
        values=values,
        
        all_columns=df.columns.tolist(),
        numeric_columns=numeric_columns,
        x=x_column,
        y=y_column,
        
        
        bar_title=bar_title,
        pie_title=pie_title,
        line_title=line_title,
        hist_title=hist_title,
        scatter_title=scatter_title,

        
        top5=top5,
        insights=insights, 
        top10=top10,
        bottom10=bottom10,

        summary=summary
    
    
    )


@app.route("/save_dashboard_charts", methods=["POST"])
def save_dashboard_charts():

    if "user" not in session:
        return redirect("/login")
    workspace_id = session.get("workspace_id")

    if not workspace_id:
        return "No workspace found.", 400

    dashboard_name = request.form.get("dashboard_name")
    x_column = request.form.get("x_column")
    y_column = request.form.get("y_column")

    # Make sure the workspace actually has a dataset
    df = get_dataframe(workspace_id)

    if df is None:
        return """
        <h2>No dataset found.</h2>
        <p>Please upload a CSV/Excel file or open a worksheet first.</p>
        <a href="/workspace">Go Back</a>
        """

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO saved_dashboards
        (
            username,
            workspace_id,
            dashboard_name,
            x_column,
            y_column
        )
        VALUES (?, ?, ?, ?, ?)
    """, (
        session["user"],
        workspace_id,
        dashboard_name,
        x_column,
        y_column
    ))

    conn.commit()
    conn.close()

    return redirect("/dashboard_charts")
@app.route("/saved_dashboards")
def saved_dashboards():

    if "user" not in session:
        return redirect("/login")

    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
    SELECT *
    FROM saved_dashboards
    WHERE username = ?
    ORDER BY id DESC
    """, (session["user"],))

    dashboards = cursor.fetchall()

    conn.close()

    return render_template(
        "saved_dashboards.html",
        dashboards=dashboards
    )


@app.route("/open_dashboard/<int:id>")
def open_dashboard(id):

    if "user" not in session:
        return redirect("/login")

    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM saved_dashboards
        WHERE id = ?
    """, (id,))

    dashboard = cursor.fetchone()

    conn.close()

    if dashboard is None:
        return "Dashboard not found"

    # Restore the saved values
    session["filename"] = dashboard["filename"]

    return redirect(
        f"/dashboard_charts?"
        f"x_column={dashboard['x_column']}"
        f"&y_column={dashboard['y_column']}"
    )


@app.route("/share_dashboard/<int:id>")
def share_dashboard(id):

    link = request.host_url + "open_dashboard/" + str(id)

    return render_template(
        "share_dashboard.html",
        link=link
    )

@app.route("/delete_dashboard/<int:id>")
def delete_dashboard(id):

    if "user" not in session:
        return redirect("/login")

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute(
        "DELETE FROM saved_dashboards WHERE id=? AND username=?",
        (id, session["user"])
    )

    conn.commit()
    conn.close()

    return redirect("/saved_dashboards")

@app.route("/edit_dashboard/<int:id>", methods=["GET", "POST"])
def edit_dashboard(id):

    if "user" not in session:
        return redirect("/login")

    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    if request.method == "POST":

        new_name = request.form.get("dashboard_name")

        cursor.execute(
            """
            UPDATE saved_dashboards
            SET dashboard_name = ?
            WHERE id = ? AND username = ?
            """,
            (new_name, id, session["user"])
        )

        conn.commit()
        conn.close()

        return redirect("/saved_dashboards")

    cursor.execute(
        """
        SELECT *
        FROM saved_dashboards
        WHERE id=? AND username=?
        """,
        (id, session["user"])
    )

    dashboard = cursor.fetchone()

    conn.close()

    return render_template(
        "edit_dashboard.html",
        dashboard=dashboard
    )


@app.route("/generate_report", methods=["GET", "POST"])
def generate_report():


    workspace_id = session.get("workspace_id")

    df = get_dataframe(workspace_id)

    if df is None:
        return """
        <h2>No dataset found.</h2>
        <p>Please upload a CSV/Excel file or open a worksheet first.</p>
        <a href="/workspace">Go Back</a>
        """

    x_column = request.form.get("x_column")
    y_column = request.form.get("y_column")
    bar_image = request.form.get("bar")
    pie_image = request.form.get("pie")
    line_image = request.form.get("line")
    histogram_image = request.form.get("histogram")
    scatter_image = request.form.get("scatter")
# Automatically select columns if the dashboard button
# did not send them
    if not x_column:
        text_columns = df.select_dtypes(include="object").columns.tolist()
        numeric_columns = df.select_dtypes(include="number").columns.tolist()

        if text_columns:
            x_column = text_columns[0]
        else:
            x_column = df.columns[0]

    if not y_column:
        numeric_columns = df.select_dtypes(include="number").columns.tolist()

        if not numeric_columns:
            return "No numeric column is available for generating the report."

        y_column = numeric_columns[0]

    total = round(df[y_column].sum(),2)
    average = round(df[y_column].mean(),2)
    maximum = df[y_column].max()
    minimum = df[y_column].min()

    top10 = (
        df[[x_column, y_column]]
        .sort_values(y_column, ascending=False)
        .head(10)
    )

    bottom10 = (
        df[[x_column, y_column]]
        .sort_values(y_column)
        .head(10)
    )

    insights = [

        f"Dataset contains {len(df)} records.",

        f"Average {y_column} is {average}.",

        f"Highest {y_column} is {maximum}.",

        f"Lowest {y_column} is {minimum}.",

        f"{x_column} contains {df[x_column].nunique()} unique values."

    ]

    if x_column == y_column:

        top10 = (
            df[[y_column]]
            .sort_values(by=y_column, ascending=False)
            .head(10)
        )

        bottom10 = (
            df[[y_column]]
            .sort_values(by=y_column, ascending=True)
            .head(10)
        )

    else:

        top10 = (
            df[[x_column, y_column]]
            .sort_values(by=y_column, ascending=False)
            .head(10)
        )

        bottom10 = (
            df[[x_column, y_column]]
            .sort_values(by=y_column, ascending=True)
            .head(10)
        )


    return render_template(
        "report.html",

        filename=session.get("filename", "Worksheet"),

        x=x_column,

        y=y_column,

        total=total,

        average=average,

        maximum=maximum,

        minimum=minimum,

        top10=top10.to_html(
            index=False,
            classes="table table-striped table-hover"
    ),

        bottom10=bottom10.to_html(
            index=False,
            classes="table table-striped table-hover"
        ),

        insights=insights
    )

def chart_image(data):

    if not data:
        return None

    image_data = data.split(",", 1)[1]

    decoded = base64.b64decode(image_data)

    return Image(
        BytesIO(decoded),
        width=6.8 * inch,
        height=3.8 * inch
    )

@app.route("/export_pdf", methods=["POST"])
def export_pdf():

    from datetime import datetime

    # Get workspace ID from form first
    workspace_id = request.form.get(
        "workspace_id",
        type=int
    )

    # If not supplied, try session
    if not workspace_id:
        workspace_id = session.get("workspace_id")

    print("PDF WORKSPACE ID:", workspace_id)

    df = get_dataframe(workspace_id)

    if df is None:
        return """
        <h2>No dataset found.</h2>
        <p>The workspace could not be loaded.</p>
        <a href="/dashboard">Back to Dashboard</a>
        """

    # Remove Unnamed columns
    df = df.loc[
        :,
        ~df.columns.astype(str).str.match(r"^Unnamed")
    ]

    # Get charts
    bar_image = request.form.get("bar")
    pie_image = request.form.get("pie")
    line_image = request.form.get("line")
    histogram_image = request.form.get("histogram")
    scatter_image = request.form.get("scatter")

    print("BAR RECEIVED:", bool(bar_image))
    print("PIE RECEIVED:", bool(pie_image))
    print("LINE RECEIVED:", bool(line_image))
    print("HISTOGRAM RECEIVED:", bool(histogram_image))
    print("SCATTER RECEIVED:", bool(scatter_image))

    if not any([
        bar_image,
        pie_image,
        line_image,
        histogram_image,
        scatter_image
    ]):

        return """
        <h2>No chart data received.</h2>
        <p>Please return to the dashboard and try again.</p>
        <a href="/dashboard">Back to Dashboard</a>
        """

    # Continue with the rest of your PDF code...

    # ==========================
    # PDF
    # ==========================

    pdf_file = "AI_Report.pdf"

    doc = SimpleDocTemplate(
        pdf_file,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()

    story = []

    # ==========================
    # REPORT INFORMATION
    # ==========================

    dataset_name = session.get(
        "filename",
        "Worksheet"
    )

    username = session.get(
        "user",
        "Unknown User"
    )

    generated_on = datetime.now().strftime(
        "%d %B %Y %I:%M %p"
    )

    # ==========================
    # TITLE
    # ==========================

    story.append(
        Paragraph(
            "<b>AI DATA ANALYSIS REPORT</b>",
            styles["Title"]
        )
    )

    story.append(Spacer(1, 15))

    story.append(
        Paragraph(
            "<b>Generated By:</b> AI Analytics Platform",
            styles["Heading2"]
        )
    )

    story.append(Spacer(1, 10))

    story.append(
        Paragraph(
            f"<b>Dataset:</b> {dataset_name}",
            styles["BodyText"]
        )
    )

    story.append(
        Paragraph(
            f"<b>User:</b> {username}",
            styles["BodyText"]
        )
    )

    story.append(
        Paragraph(
            f"<b>Generated On:</b> {generated_on}",
            styles["BodyText"]
        )
    )

    story.append(Spacer(1, 20))

    # ==========================
    # INTRODUCTION
    # ==========================

    story.append(
        Paragraph(
            "This report was automatically generated by the "
            "AI Analytics Platform. It summarizes the dataset "
            "and provides useful insights and visualizations.",
            styles["BodyText"]
        )
    )

    story.append(Spacer(1, 20))

    # ==========================
    # DATASET OVERVIEW
    # ==========================

    story.append(
        Paragraph(
            "<b>DATASET OVERVIEW</b>",
            styles["Heading1"]
        )
    )

    missing = int(df.isnull().sum().sum())
    duplicates = int(df.duplicated().sum())

    story.append(
        Paragraph(
            f"Total Rows: {len(df)}",
            styles["BodyText"]
        )
    )

    story.append(
        Paragraph(
            f"Total Columns: {len(df.columns)}",
            styles["BodyText"]
        )
    )

    story.append(
        Paragraph(
            f"Missing Values: {missing}",
            styles["BodyText"]
        )
    )

    story.append(
        Paragraph(
            f"Duplicate Rows: {duplicates}",
            styles["BodyText"]
        )
    )

    story.append(Spacer(1, 20))

    # ==========================
    # KPI
    # ==========================

    story.append(
        Paragraph(
            "<b>KPI SUMMARY</b>",
            styles["Heading1"]
        )
    )

    kpi_data = [
        ["Metric", "Value"],
        ["Total Rows", len(df)],
        ["Total Columns", len(df.columns)],
        ["Missing Values", missing],
        ["Duplicate Rows", duplicates]
    ]

    numeric_columns = df.select_dtypes(
        include="number"
    ).columns

    for col in numeric_columns:

        kpi_data.append([
            f"Total {col}",
            round(df[col].sum(), 2)
        ])

        kpi_data.append([
            f"Average {col}",
            round(df[col].mean(), 2)
        ])

        kpi_data.append([
            f"Maximum {col}",
            df[col].max()
        ])

        kpi_data.append([
            f"Minimum {col}",
            df[col].min()
        ])

        kpi_data.append([
            f"Median {col}",
            round(df[col].median(), 2)
        ])

    kpi_table = Table(
        kpi_data,
        repeatRows=1
    )

    kpi_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10)
        ])
    )

    story.append(kpi_table)

    # ==========================
    # AI INSIGHTS
    # ==========================

    story.append(Spacer(1, 20))

    story.append(
        Paragraph(
            "<b>AI INSIGHTS</b>",
            styles["Heading1"]
        )
    )

    insights = []

    insights.append(
        f"The dataset contains {len(df)} rows and "
        f"{len(df.columns)} columns."
    )

    if missing == 0:
        insights.append(
            "No missing values were detected."
        )
    else:
        insights.append(
            f"The dataset contains {missing} missing values."
        )

    if duplicates == 0:
        insights.append(
            "No duplicate rows were found."
        )
    else:
        insights.append(
            f"There are {duplicates} duplicate rows."
        )

    numeric = df.select_dtypes(include="number")

    if not numeric.empty:

        highest_avg = numeric.mean().idxmax()

        highest_value = numeric.mean().max()

        insights.append(
            f"'{highest_avg}' has the highest average "
            f"value of {highest_value:.2f}."
        )

    for item in insights:

        story.append(
            Paragraph(
                f"• {item}",
                styles["BodyText"]
            )
        )

    # ==========================
    # TOP RECORDS
    # ==========================

    story.append(Spacer(1, 20))

    story.append(
        Paragraph(
            "<b>TOP 10 DATA RECORDS</b>",
            styles["Heading1"]
        )
    )

    preview_df = df.head(10)

    table_data = [
        [str(col) for col in preview_df.columns]
    ]

    for row in preview_df.values.tolist():

        table_data.append([
            str(value)
            for value in row
        ])

    preview_table = Table(
        table_data,
        repeatRows=1
    )

    preview_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER")
        ])
    )

    story.append(preview_table)

    # ==========================
    # CHART FUNCTION
    # ==========================

    def add_chart(title, image_data):

        if not image_data:
            return

        try:

            # Remove data:image/png;base64,
            if "," in image_data:
                image_data = image_data.split(",", 1)[1]

            image_bytes = base64.b64decode(
                image_data
            )

            image_stream = io.BytesIO(
                image_bytes
            )

            img = Image(
                image_stream,
                width=500,
                height=280
            )

            story.append(
                Paragraph(
                    f"<b>{title}</b>",
                    styles["Heading2"]
                )
            )

            story.append(img)

            story.append(
                Spacer(1, 20)
            )

        except Exception as e:

            print(
                f"Error loading {title}:",
                e
            )

    # ==========================
    # DATA VISUALIZATIONS
    # ==========================

    story.append(PageBreak())

    story.append(
        Paragraph(
            "<b>DATA VISUALIZATIONS</b>",
            styles["Heading1"]
        )
    )

    story.append(Spacer(1, 15))

    add_chart(
        "Bar Chart",
        bar_image
    )

    add_chart(
        "Pie Chart",
        pie_image
    )

    story.append(PageBreak())

    add_chart(
        "Line Chart",
        line_image
    )

    add_chart(
        "Histogram",
        histogram_image
    )

    add_chart(
        "Scatter Plot",
        scatter_image
    )

    # ==========================
    # BUILD
    # ==========================

    doc.build(story)

    return send_file(
        pdf_file,
        as_attachment=True,
        download_name="AI_Data_Analysis_Report.pdf"
    )
@app.route("/export_word")
def export_word():

    document = Document()

    document.add_heading("AI Data Analysis Report", level=1)

    document.add_paragraph(
        "Generated by AI Analytics Platform"
    )

    document.add_heading("Dataset", level=2)

    document.add_paragraph(session["filename"])

    document.add_heading("Summary", level=2)

    document.add_paragraph(
        "This report contains automated business insights."
    )

    document.add_heading("Generated On", level=2)

    from datetime import datetime

    document.add_paragraph(
        datetime.now().strftime("%d %B %Y %H:%M")
    )

    filename = "AI_Report.docx"

    document.save(filename)

    return send_file(
        filename,
        as_attachment=True
    )


@app.route("/export_powerpoint")
def export_powerpoint():

    prs = Presentation()

    # -------------------------
    # Slide 1 - Cover
    # -------------------------
    slide_layout = prs.slide_layouts[0]
    slide = prs.slides.add_slide(slide_layout)

    slide.shapes.title.text = "AI Data Analysis Report"

    slide.placeholders[1].text = (
        f"Dataset: {session['filename']}\n"
        f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}"
    )

    # -------------------------
    # Slide 2 - KPI Summary
    # -------------------------
    slide_layout = prs.slide_layouts[5]
    slide = prs.slides.add_slide(slide_layout)

    slide.shapes.title.text = "Dashboard Summary"

    textbox = slide.shapes.add_textbox(
        Inches(0.7),
        Inches(1),
        Inches(8),
        Inches(4)
    )

    tf = textbox.text_frame

    tf.text = f"Rows: {dashboard['rows']}"

    tf.add_paragraph().text = f"Columns: {dashboard['columns']}"

    tf.add_paragraph().text = f"Total: {dashboard['total']}"

    tf.add_paragraph().text = f"Average: {dashboard['average']}"

    tf.add_paragraph().text = f"Maximum: {dashboard['maximum']}"

    tf.add_paragraph().text = f"Minimum: {dashboard['minimum']}"

    # -------------------------
    # Slide 3 - Insights
    # -------------------------
    slide = prs.slides.add_slide(prs.slide_layouts[5])

    slide.shapes.title.text = "AI Insights"

    box = slide.shapes.add_textbox(
        Inches(0.6),
        Inches(1),
        Inches(8),
        Inches(5)
    )

    frame = box.text_frame

    for item in insights:
        frame.add_paragraph().text = item

    # -------------------------
    # Save
    # -------------------------
    filename = "AI_Report.pptx"

    prs.save(filename)

    return send_file(
        filename,
        as_attachment=True
    )


@app.route("/new_workspace")
def new_workspace():

    return render_template("new_workspace.html")
@app.route("/create_workspace", methods=["POST"])
def create_workspace():

    if "user" not in session:
        return redirect("/login")

    workspace_name = request.form["workspace_name"]
    workspace_type = request.form["workspace_type"]

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""

    INSERT INTO workspaces(

        username,
        workspace_name,
        workspace_type,
        worksheet_data

    )

    VALUES(?,?,?,?)

    """,(

        session["user"],
        workspace_name,
        workspace_type,
        "[]"

    ))

    conn.commit()

    workspace_id = cursor.lastrowid

    conn.close()

    return redirect(f"/worksheet/{workspace_id}")
WORKSPACE_TEMPLATES = {

    "school": [
        "Student Name",
        "Class",
        "English",
        "Mathematics",
        "Science"
    ],

    "business": [
        "Date",
        "Product",
        "Quantity",
        "Price",
        "Customer"
    ],

    "church": [
        "Member",
        "Department",
        "Attendance",
        "Offering",
        "Date"
    ],

    "inventory": [
        "Product",
        "Quantity",
        "Cost Price",
        "Selling Price",
        "Supplier"
    ],

    "farm": [
        "Date",
        "Crop",
        "Harvest",
        "Price",
        "Location"
    ]

}
@app.route("/worksheet/<int:workspace_id>")
def worksheet(workspace_id):

    if "user" not in session:
        return redirect("/login")

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT workspace_name,
               workspace_type,
               worksheet_data
        FROM workspaces
        WHERE id = ?
        AND username = ?
    """, (
        workspace_id,
        session["user"]
    ))

    workspace = cursor.fetchone()

    conn.close()

    if not workspace:
        return "Workspace not found."

    workspace_name = workspace[0]
    workspace_type = workspace[1]

    # Make this the active workspace
    session["workspace_id"] = workspace_id

    # Load worksheet data
    if workspace[2]:
        worksheet_data = json.loads(workspace[2])
    else:
        worksheet_data = None

    print("WORKSHEET ID:", workspace_id)
    print("Raw database value:", workspace[2])
    print("Loaded worksheet data:", worksheet_data)

    columns = WORKSPACE_TEMPLATES.get(
        workspace_type,
        ["Column 1", "Column 2", "Column 3"]
    )

    return render_template(
        "worksheet.html",
        workspace_id=workspace_id,
        workspace_name=workspace_name,
        workspace_type=workspace_type,
        columns=columns,
        worksheet_data=worksheet_data
    )

@app.route("/save_worksheet", methods=["POST"])
def save_worksheet():

    try:

        print("Save request received")

        if "user" not in session:
            return jsonify({"message": "Login required"}), 401

        data = request.get_json()

        print(data)

        workspace_id = data["workspace_id"]
        worksheet = json.dumps(data["data"])

        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE workspaces
        SET worksheet_data=?
        WHERE id=? AND username=?
        """, (
            worksheet,
            workspace_id,
            session["user"]
        ))

        print("Rows updated:", cursor.rowcount)

        conn.commit()
        conn.close()

        return jsonify({
            "message": "Worksheet saved successfully."
        })

    except Exception as e:
        print("ERROR:", e)
        return jsonify({"message": str(e)}), 500
if __name__ == "__main__":
    app.run(debug=True)                                                                                                                                                                                        
                                                                                                                                                                               